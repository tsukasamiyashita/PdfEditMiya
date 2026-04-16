# -*- coding: utf-8 -*-
import os, sys, cv2, csv, json, time, gc, re
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
import fitz  # PyMuPDF
import pytesseract
import ezdxf
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from PIL import Image

try:
    from docx import Document
except ImportError:
    Document = None

try:
    import xlrd
except ImportError:
    xlrd = None

from common import (
    auto_adjust_excel_column_width,
    sanitize_excel_text,
    merge_2d_arrays_horizontally,
    parse_row_data,
    apply_text_inheritance
)

# ==============================
# 画像前処理タスク (OCR精度向上・クロップ拡張)
# ==============================
def expand_crop_rect_for_intersecting_objects(img_array, rx1, ry1, rx2, ry2):
    """
    指定された相対座標(rx1, ry1, rx2, ry2)のクロップ枠に対し、
    枠の境界に接している文字や図形（輪郭）が途切れないよう、
    輪郭が完全に含まれるようにピクセル座標レベルで枠を自動拡張して返す。
    水平に近い線（高さが極小）の場合は、左右(X)の拡張を抑え、上下(Y)の拡張を優先する。
    """
    h, w = img_array.shape[:2]
    x1, y1 = int(min(rx1, rx2) * w), int(min(ry1, ry2) * h)
    x2, y2 = int(max(rx1, rx2) * w), int(max(ry1, ry2) * h)
    
    # なぞり方向の判定
    is_v_line = abs(rx2 - rx1) < 0.03 and abs(ry2 - ry1) > abs(rx2 - rx1)
    is_h_line = abs(ry2 - ry1) < 0.03 and not is_v_line

    if x1 == 0 and y1 == 0 and x2 == w and y2 == h:
        return x1, y1, x2, y2

    if len(img_array.shape) == 3:
        gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
    else:
        gray = img_array.copy()

    clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
    gray = clahe.apply(gray)
    gray = cv2.medianBlur(gray, 3)

    binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2)
    contours, _ = cv2.findContours(binary, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    
    new_x1, new_y1, new_x2, new_y2 = x1, y1, x2, y2
    
    for cnt in contours:
        cx, cy, cw, ch = cv2.boundingRect(cnt)
        cx1, cy1, cx2, cy2 = cx, cy, cx + cw, cy + ch
        
        if cw > w * 0.5 or ch > h * 0.5: continue
        if cw < 3 or ch < 3: continue
            
        if (cx1 <= x2 and cx2 >= x1 and cy1 <= y2 and cy2 >= y1):
            if is_h_line:
                # 水平線モード：ユーザーの引いた「水平線（中心Y）」が文字の矩形と重なっている場合のみ
                target_y = (y1 + y2) / 2
                expand_needed = (cy1 <= target_y <= cy2)
                if expand_needed:
                    intersect_x1, intersect_x2 = max(cx1, x1), min(cx2, x2)
                    if (intersect_x2 - intersect_x1) < cw * 0.5: expand_needed = False
            elif is_v_line:
                # 垂直線モード：ユーザーの引いた「垂直線（中心X）」が文字の矩形と重なっている場合のみ
                target_x = (x1 + x2) / 2
                expand_needed = (cx1 <= target_x <= cx2)
                if expand_needed:
                    intersect_y1, intersect_y2 = max(cy1, y1), min(cy2, y2)
                    if (intersect_y2 - intersect_y1) < ch * 0.5: expand_needed = False
            else:
                expand_needed = True

            if expand_needed:
                if not is_v_line:
                    new_x1 = min(new_x1, cx1)
                    new_x2 = max(new_x2, cx2)
                if not is_h_line:
                    new_y1 = min(new_y1, cy1)
                    new_y2 = max(new_y2, cy2)
            
    margin = 4
    new_x1 = max(0, new_x1 - (0 if is_v_line else margin))
    new_y1 = max(0, new_y1 - (0 if is_h_line else margin))
    new_x2 = min(w, new_x2 + (0 if is_v_line else margin))
    new_y2 = min(h, new_y2 + (0 if is_h_line else margin))
    
    return (x1 if is_v_line else new_x1), (y1 if is_h_line else new_y1), (x2 if is_v_line else new_x2), (y2 if is_h_line else new_y2)

# ==============================
# コア処理関数群 (PDF編集・標準抽出)
# ==============================
def check_pdf_has_text(pdf_path):
    """PDFにテキストデータ（フォント情報）が含まれているか確認する"""
    try:
        with fitz.open(pdf_path) as doc:
            for page in doc:
                if page.get_text().strip():
                    return True
    except Exception:
        pass
    return False

def merge_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    writer = PdfWriter()
    ui.update_overall(0, len(files), f"全体の進捗 ( 0 / {len(files)} ファイル )")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ファイルを結合中... ( {j} / {len(reader.pages)} ページ )")
            writer.add_page(p)
    ui.set_determinate(1, 1, "PDFを保存中...")
    if save_dir:
        with open(os.path.join(save_dir, f"{options.get('folder_name', 'Merged')}_Merge.pdf"), "wb") as out: writer.write(out)

def split_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(reader.pages))))
        for n, p in enumerate(reader.pages, 1):
            ui.set_determinate(n, len(reader.pages), f"ファイルを分割中... ( {n} / {len(reader.pages)} ページ )")
            writer = PdfWriter(); writer.add_page(p)
            with open(os.path.join(save_dir, f"{base}_Split_{str(n).zfill(digits)}.pdf"), "wb") as out: writer.write(out)

def rotate_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f); writer = PdfWriter()
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ページを回転中... ( {j} / {len(reader.pages)} ページ )")
            p.rotate(options.get("rotate_deg", 270)); writer.add_page(p)
        base = os.path.splitext(os.path.basename(f))[0]
        with open(os.path.join(save_dir, f"{base}_Rotate.pdf"), "wb") as out: writer.write(out)

def extract_text_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    
    for f in files:
        if not check_pdf_has_text(f):
            raise Exception(f"ファイル「{os.path.basename(f)}」はスキャンされた画像（ラスターデータ）のため、標準ライブラリでは文字を抽出できません。\nエンジンを「Gemini API」または「Tesseract」に変更して実行してください。")
            
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(f))[0]; text_list = []
        is_scanned_pdf = False
        with pdfplumber.open(f) as pdf:
            for j, p in enumerate(pdf.pages, 1):
                ui.set_determinate(j, len(pdf.pages), f"テキストを抽出中... ( {j} / {len(pdf.pages)} ページ )")
                if not p.chars: is_scanned_pdf = True
                
                if crop_regions:
                    page_texts = []
                    for idx, region in enumerate(crop_regions, 1):
                        rx1, ry1, rx2, ry2 = region[:4]
                        is_vert = region[4] if len(region) > 4 else False
                        # 水平・垂直線モードの判定（明示的なフラグがない場合のフォールバックも含む）
                        is_line = is_vert or abs(ry2 - ry1) < 0.03 or abs(rx2 - rx1) < 0.03
                        
                        m_y = 0.0 if (is_line and is_vert) else 0.005
                        m_x = 0.0 if (is_line and not is_vert) else 0.005
                        
                        x0 = max(0, min(rx1, rx2) - m_x) * p.width
                        top = max(0, min(ry1, ry2) - m_y) * p.height
                        x1 = min(1, max(rx1, rx2) + m_x) * p.width
                        bottom = min(1, max(ry1, ry2) + m_y) * p.height
                        
                        cropped_page = p.crop((x0, top, x1, bottom), strict=False)
                        
                        if is_line:
                            if is_vert:
                                target_x = (min(rx1, rx2) + max(rx1, rx2)) / 2 * p.width
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_y = (obj["top"] + obj["bottom"]) / 2
                                        intersects_x = (obj["x0"] <= target_x <= obj["x1"])
                                        return (top <= mid_y <= bottom) and intersects_x
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)
                            else:
                                target_y = (min(ry1, ry2) + max(ry1, ry2)) / 2 * p.height
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_x = (obj["x0"] + obj["x1"]) / 2
                                        intersects_y = (obj["top"] <= target_y <= obj["bottom"])
                                        return (x0 <= mid_x <= x1) and intersects_y
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)

                        txt = cropped_page.extract_text(layout=True)
                        if txt: page_texts.append(f"--- 範囲{idx} ---\n{txt}")
                    if page_texts: text_list.append(f"【Page {j}】\n" + "\n".join(page_texts))
                else:
                    txt = p.extract_text()
                    if txt: text_list.append(f"【Page {j}】\n{txt}")
        
        if text_list:
            output_content = "\n\n".join(text_list)
        elif is_scanned_pdf:
            output_content = "このPDFは「画像（スキャンされたPDF）」として保存されており、標準ライブラリでは文字を読み取れません。\n「Gemini API」または「Tesseract」エンジンを使用して再試行してください。"
        else:
            output_content = "指定された範囲内にテキストデータが見つかりませんでした。枠を少し広げて選択するか、AIエンジンの使用を検討してください。"
            
        with open(os.path.join(save_dir, f"{base}_Text.txt"), "w", encoding="utf-8") as out: out.write(output_content)

def convert_to_excel_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    
    for f in files:
        if not check_pdf_has_text(f):
            raise Exception(f"ファイル「{os.path.basename(f)}」はスキャンされた画像（ラスターデータ）のため、標準ライブラリではデータを抽出できません。\nエンジンを「Gemini API」または「Tesseract」に変更して実行してください。")
            
    border_style = Side(border_style="thin", color="000000")
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        wb = Workbook(); wb.remove(wb.active)
        base_name = os.path.splitext(os.path.basename(pdf_path))[0]
        
        has_any_table = False
        is_scanned_pdf = False
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"表データを抽出中... ( {page_idx} / {len(pdf.pages)} ページ )")
                if not page.chars: is_scanned_pdf = True
                
                tables = []
                if crop_regions:
                    all_regions_data = []
                    for region in crop_regions:
                        rx1, ry1, rx2, ry2 = region[:4]
                        is_vert = region[4] if len(region) > 4 else False
                        is_line = is_vert or abs(ry2 - ry1) < 0.03 or abs(rx2 - rx1) < 0.03
                        
                        m_y = 0.0 if (is_line and is_vert) else 0.005
                        m_x = 0.0 if (is_line and not is_vert) else 0.005
                        
                        x0 = max(0, min(rx1, rx2) - m_x) * page.width
                        top = max(0, min(ry1, ry2) - m_y) * page.height
                        x1 = min(1, max(rx1, rx2) + m_x) * page.width
                        bottom = min(1, max(ry1, ry2) + m_y) * page.height
                        
                        cropped_page = page.crop((x0, top, x1, bottom), strict=False)
                        
                        if is_line:
                            if is_vert:
                                target_x = (min(rx1, rx2) + max(rx1, rx2)) / 2 * page.width
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_y = (obj["top"] + obj["bottom"]) / 2
                                        intersects_x = (obj["x0"] <= target_x <= obj["x1"])
                                        return (top <= mid_y <= bottom) and intersects_x
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)
                            else:
                                target_y = (min(ry1, ry2) + max(ry1, ry2)) / 2 * page.height
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_x = (obj["x0"] + obj["x1"]) / 2
                                        intersects_y = (obj["top"] <= target_y <= obj["bottom"])
                                        return (x0 <= mid_x <= x1) and intersects_y
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)
                        
                        tbls = []
                        if options.get("extract_mode") != "text":
                            tbls = cropped_page.extract_tables()
                            if not tbls:
                                tbl_settings = {"vertical_strategy": "text", "horizontal_strategy": "text", "snap_tolerance": 3}
                                tbls = cropped_page.extract_tables(table_settings=tbl_settings)
                        
                        if not tbls:
                            txt = cropped_page.extract_text()
                            if txt and txt.strip():
                                lines = [line.strip() for line in txt.strip().split('\n') if line.strip()]
                                dummy_table = [[line] for line in lines]
                                tbls = [dummy_table]
                                
                        region_table = []
                        if tbls:
                            for tbl in tbls: region_table.extend(tbl)
                        
                        if not region_table: region_table = [[""]]
                        all_regions_data.append(region_table)
                        
                    if options.get("extract_mode") == "text":
                        merged_row = []
                        for region_data in all_regions_data:
                            region_texts = []
                            for r in region_data:
                                for c in r:
                                    if str(c).strip(): region_texts.append(str(c).strip())
                            merged_row.append("\n".join(region_texts) if region_texts else "")
                        merged_table = [merged_row] if merged_row else [[""]]
                    else:
                        merged_table = merge_2d_arrays_horizontally(all_regions_data)
                        
                    is_empty = True
                    for row in merged_table:
                        for cell in row:
                            if cell and str(cell).strip():
                                is_empty = False; break
                        if not is_empty: break
                            
                    if not is_empty: tables = [merged_table]
                else: 
                    tables = []
                    if options.get("extract_mode") != "text":
                        tables = page.extract_tables()
                        
                    if not tables:
                        txt = page.extract_text()
                        if txt and txt.strip():
                            lines = [line.strip() for line in txt.strip().split('\n') if line.strip()]
                            if options.get("extract_mode") == "text":
                                tables = [[lines]]
                            else:
                                dummy_table = [[line] for line in lines]
                                tables = [dummy_table]
                
                if not tables: continue
                has_any_table = True
                ws = wb.create_sheet(f"Page_{str(page_idx).zfill(digits)}"); current_row = 1
                for table in tables:
                    for row_data in table:
                        for col_idx, cell_value in enumerate(row_data, 1):
                            clean_value = sanitize_excel_text(cell_value)
                            cell = ws.cell(row=current_row, column=col_idx, value=clean_value if clean_value else "")
                            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                        current_row += 1
                    current_row += 2
                auto_adjust_excel_column_width(ws)
        
        if has_any_table:
            wb.save(os.path.join(save_dir, f"{base_name}_Excel.xlsx"))
        else:
            empty_wb = Workbook(); ws = empty_wb.active; ws.title = "No_Data"
            if is_scanned_pdf: msg = "このPDFは「スキャンされた画像」です。Gemini APIまたはTesseractを使用してください。"
            else: msg = "指定範囲内に表構造やテキストデータが見つかりませんでした。"
            ws.cell(row=1, column=1, value=msg)
            empty_wb.save(os.path.join(save_dir, f"{base_name}_Excel_NoData.xlsx"))

def convert_to_csv_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    
    for f in files:
        if not check_pdf_has_text(f):
            raise Exception(f"ファイル「{os.path.basename(f)}」はスキャンされた画像（ラスターデータ）のため、標準ライブラリではデータを抽出できません。\nエンジンを「Gemini API」または「Tesseract」に変更して実行してください。")
            
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        
        has_any_table = False
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"CSV変換中... ( {page_idx} / {len(pdf.pages)} ページ )")
                tables = []
                if crop_regions:
                    all_regions_data = []
                    for region in crop_regions:
                        rx1, ry1, rx2, ry2 = region[:4]
                        is_vert = region[4] if len(region) > 4 else False
                        is_line = is_vert or abs(ry2 - ry1) < 0.03 or abs(rx2 - rx1) < 0.03
                        
                        m_y = 0.0 if (is_line and is_vert) else 0.005
                        m_x = 0.0 if (is_line and not is_vert) else 0.005
                        
                        x0 = max(0, min(rx1, rx2) - m_x) * page.width
                        top = max(0, min(ry1, ry2) - m_y) * page.height
                        x1 = min(1, max(rx1, rx2) + m_x) * page.width
                        bottom = min(1, max(ry1, ry2) + m_y) * page.height
                        
                        cropped_page = page.crop((x0, top, x1, bottom), strict=False)
                        
                        if is_line:
                            if is_vert:
                                target_x = (min(rx1, rx2) + max(rx1, rx2)) / 2 * page.width
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_y = (obj["top"] + obj["bottom"]) / 2
                                        intersects_x = (obj["x0"] <= target_x <= obj["x1"])
                                        return (top <= mid_y <= bottom) and intersects_x
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)
                            else:
                                target_y = (min(ry1, ry2) + max(ry1, ry2)) / 2 * page.height
                                def filter_obj(obj):
                                    if obj.get("object_type") == "char":
                                        mid_x = (obj["x0"] + obj["x1"]) / 2
                                        intersects_y = (obj["top"] <= target_y <= obj["bottom"])
                                        return (x0 <= mid_x <= x1) and intersects_y
                                    return True
                                cropped_page = cropped_page.filter(filter_obj)
                        
                        tbls = []
                        if options.get("extract_mode") != "text":
                            tbls = cropped_page.extract_tables()
                            if not tbls:
                                tbls = cropped_page.extract_tables(table_settings={"vertical_strategy": "text", "horizontal_strategy": "text"})
                        
                        if not tbls:
                            txt = cropped_page.extract_text()
                            if txt and txt.strip():
                                lines = [line.strip() for line in txt.strip().split('\n') if line.strip()]
                                tbls = [[[line] for line in lines]]
                                
                        region_table = []
                        if tbls:
                            for tbl in tbls: region_table.extend(tbl)
                                
                        if not region_table: region_table = [[""]]
                        all_regions_data.append(region_table)
                        
                    if options.get("extract_mode") == "text":
                        merged_row = []
                        for region_data in all_regions_data:
                            region_texts = []
                            for r in region_data:
                                for c in r:
                                    if str(c).strip(): region_texts.append(str(c).strip())
                            merged_row.append("\n".join(region_texts) if region_texts else "")
                        merged_table = [merged_row] if merged_row else [[""]]
                    else:
                        merged_table = merge_2d_arrays_horizontally(all_regions_data)
                        
                    is_empty = True
                    for row in merged_table:
                        for cell in row:
                            if cell and str(cell).strip():
                                is_empty = False; break
                        if not is_empty: break
                            
                    if not is_empty: tables = [merged_table]
                else: 
                    tables = []
                    if options.get("extract_mode") != "text":
                        tables = page.extract_tables()
                        
                    if not tables:
                        txt = page.extract_text()
                        if txt and txt.strip():
                            lines = [line.strip() for line in txt.strip().split('\n') if line.strip()]
                            if options.get("extract_mode") == "text":
                                tables = [[lines]]
                            else:
                                tables = [[[line] for line in lines]]
                
                if not tables: continue
                has_any_table = True
                with open(os.path.join(save_dir, f"{base}_Page_{str(page_idx).zfill(digits)}_CSV.csv"), "w", encoding="utf-8-sig", newline="") as f_out:
                    writer = csv.writer(f_out)
                    for table in tables:
                        for row_data in table: writer.writerow([str(cell).strip() if cell else "" for cell in row_data])
                        writer.writerow([]) 
        
        if not has_any_table:
            with open(os.path.join(save_dir, f"{base}_NoData_CSV.txt"), "w", encoding="utf-8") as f_out:
                f_out.write("データが見つかりませんでした。PDFがスキャン形式であるか、範囲が狭すぎる可能性があります。")

def convert_to_image_jpg(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "jpg")
def convert_to_image_png(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "png")
def convert_to_image_tiff(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "tiff")
def convert_to_image_bmp(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "bmp")

def _convert_image(files, save_dir, options, ui, ext):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        doc = fitz.open(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(doc))))
        for n, page in enumerate(doc, 1):
            ui.set_determinate(n, len(doc), f"画像へ変換中... ( {n} / {len(doc)} ページ )")
            n_str = str(n).zfill(digits)
            pix = page.get_pixmap(dpi=200)
            if crop_regions:
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
                elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)
                h, w = img_array.shape[:2]
                for idx, (rx1, ry1, rx2, ry2) in enumerate(crop_regions):
                    x1, y1 = int(min(rx1, rx2) * w), int(min(ry1, ry2) * h)
                    x2, y2 = int(max(rx1, rx2) * w), int(max(ry1, ry2) * h)
                    Image.fromarray(img_array[y1:y2, x1:x2]).save(os.path.join(save_dir, f"{base}_{n_str}_crop{idx+1}.{ext}"))
            else: 
                if ext in ["tiff", "bmp"]:
                    Image.frombytes("RGB" if pix.n >= 3 else "L", [pix.width, pix.height], pix.samples).save(os.path.join(save_dir, f"{base}_{n_str}.{ext}"))
                else:
                    pix.save(os.path.join(save_dir, f"{base}_{n_str}.{ext}"))

def convert_to_svg(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        doc = fitz.open(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(doc))))
        for n, page in enumerate(doc, 1):
            ui.set_determinate(n, len(doc), f"SVGへ変換中... ( {n} / {len(doc)} ページ )")
            n_str = str(n).zfill(digits)
            if crop_regions:
                h, w = page.rect.height, page.rect.width
                for idx, (rx1, ry1, rx2, ry2) in enumerate(crop_regions):
                    rect = fitz.Rect(min(rx1, rx2)*w, min(ry1, ry2)*h, max(rx1, rx2)*w, max(ry1, ry2)*h)
                    page.set_cropbox(rect)
                    svg_xml = page.get_svg_image()
                    with open(os.path.join(save_dir, f"{base}_{n_str}_crop{idx+1}.svg"), "w", encoding="utf-8") as svg_file:
                        svg_file.write(svg_xml)
                    page.set_cropbox(page.mediabox) 
            else:
                svg_xml = page.get_svg_image()
                with open(os.path.join(save_dir, f"{base}_{n_str}.svg"), "w", encoding="utf-8") as svg_file:
                    svg_file.write(svg_xml)

def convert_to_dxf(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    def in_crop(x, y, w, h):
        if not crop_regions: return True
        for (rx1, ry1, rx2, ry2) in crop_regions:
            if min(rx1, rx2)*w <= x <= max(rx1, rx2)*w and min(ry1, ry2)*h <= y <= max(ry1, ry2)*h: return True
        return False

    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        try:
            doc = fitz.open(f); dwg = ezdxf.new('R2010'); msp = dwg.modelspace()
            for page_num, page in enumerate(doc, 1):
                ui.set_determinate(page_num, len(doc), f"DXFへ変換中... ( {page_num} / {len(doc)} ページ )")
                h, w = page.rect.height, page.rect.width
                paths = page.get_drawings()
                is_vector_rich = len(paths) > 0
                if is_vector_rich:
                    for path in paths:
                        for item in path["items"]:
                            if item[0] == "l":
                                p1, p2 = item[1], item[2]
                                if in_crop(p1.x, p1.y, w, h) or in_crop(p2.x, p2.y, w, h): msp.add_line((p1.x, h - p1.y), (p2.x, h - p2.y))
                            elif item[0] == "re":
                                rect = item[1]
                                if in_crop(rect.x0, rect.y0, w, h) or in_crop(rect.x1, rect.y1, w, h):
                                    msp.add_lwpolyline([(rect.x0, h - rect.y0), (rect.x1, h - rect.y0), (rect.x1, h - rect.y1), (rect.x0, h - rect.y1)], close=True)
                            elif item[0] == "c":
                                p1, p2, p3, p4 = item[1], item[2], item[3], item[4]
                                if any(in_crop(pt.x, pt.y, w, h) for pt in [p1, p2, p3, p4]):
                                    msp.add_spline([(p1.x, h - p1.y), (p2.x, h - p2.y), (p3.x, h - p3.y), (p4.x, h - p4.y)])
            ui.set_determinate(len(doc), len(doc), "DXFファイルを保存中...")
            dwg.saveas(os.path.join(save_dir, f"{os.path.splitext(os.path.basename(f))[0]}_CAD.dxf"))
        except Exception as e: print(f"DXF Conversion Error: {e}")

# ==============================
# 画像前処理タスク (OCR精度向上)
# ==============================
def preprocess_image_for_ocr(img_array, is_digital=False):
    """
    デジタルPDFとスキャンPDFで適切な前処理を自動で使い分ける
    """
    if len(img_array.shape) == 3: gray = cv2.cvtColor(img_array, cv2.COLOR_RGB2GRAY)
    else: gray = img_array
    
    if is_digital:
        # デジタル生成PDF（Excel等からの変換）の場合
        # 元々クリアな文字なので、過剰な補正(適応的二値化や平滑化)を行うと逆に文字が劣化します。
        # 大津の二値化のみを適用し、くっきりとした白黒画像にします。
        _, binary = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY | cv2.THRESH_OTSU)
        return binary
    else:
        # スキャンされたPDFの場合
        # 影や照明ムラを取り除くため、コントラストを強調してから適応的二値化を行います。
        clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8))
        gray = clahe.apply(gray)
        # 解像度が高い(dpi=400)場合、文字の線が太くなるため、ブロックサイズを大きく(51)して太い文字の中抜けを防止します。
        binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 51, 15)
        return binary

# ==============================
# ローカルOCR抽出タスク (Tesseract)
# ==============================
def check_tesseract_installation():
    if sys.platform.startswith("win"):
        tess_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(tess_path): pytesseract.pytesseract.tesseract_cmd = tess_path
    try: pytesseract.get_tesseract_version()
    except Exception: raise Exception("Tesseract OCRが見つかりません。")

def extract_tesseract_task(files, save_dir, options, ui):
    check_tesseract_installation()
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFが含まれていません。")
    out_format = options.get("out_format", "xlsx")
    crop_regions = options.get("crop_regions", [])
    
    for i, f in enumerate(files, 1):
        if ui.is_cancelled(): return
        ui.update_overall(i, len(files), f"全体進捗 ( {i} / {len(files)} )")
        base = os.path.splitext(os.path.basename(f))[0]
        doc = fitz.open(f)
        total_pages = len(doc)
        for page_num in range(total_pages):
            if ui.is_cancelled(): return
            ui.set_indeterminate(f"OCR解析中... ({page_num+1}/{total_pages}ページ)")
            
            # 【改善】PDFページ自体にテキストデータが埋め込まれているか(デジタルPDFか)を自動判定
            page_obj = doc[page_num]
            is_digital = bool(page_obj.get_text().strip())
            
            # 【改善】DPIを300から400に引き上げ、Tesseractの認識精度を向上（文字が小さい表などに効果絶大）
            pix = page_obj.get_pixmap(dpi=400)
            img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
            if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
            elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)
            
            cropped_images = []
            if crop_regions:
                h_img, w_img = img_array.shape[:2]
                for region in crop_regions:
                    rx1, ry1, rx2, ry2 = region[:4]
                    is_vert = region[4] if len(region) > 4 else False
                    is_line = is_vert or abs(ry2 - ry1) < 0.03 or abs(rx2 - rx1) < 0.03
                    # 水平線モード、垂直線モード、またはテキスト抽出モードの場合は範囲拡張
                    if out_format not in ["xlsx", "csv"] or is_line or options.get("extract_mode") == "text":
                        x1, y1, x2, y2 = expand_crop_rect_for_intersecting_objects(img_array, rx1, ry1, rx2, ry2)
                    else:
                        x1, y1 = int(min(rx1, rx2) * w_img), int(min(ry1, ry2) * h_img)
                        x2, y2 = int(max(rx1, rx2) * w_img), int(max(ry1, ry2) * h_img)
                    cropped_images.append(img_array[y1:y2, x1:x2])
            else: cropped_images.append(img_array)
                
            all_regions_data = []
            for crop_img in cropped_images:
                try:
                    # デジタルかスキャンかを渡して、最適な前処理を適用する
                    processed_img = preprocess_image_for_ocr(crop_img, is_digital)
                    
                    # 【改善】範囲指定抽出(セル単位等)の場合は PSM 6 (均一なテキストブロック)、ページ全体の場合は PSM 3 を使用
                    psm_val = 6 if crop_regions else 3
                    custom_config = f'--oem 3 --psm {psm_val}'
                    
                    text = pytesseract.image_to_string(Image.fromarray(processed_img), lang="jpn+jpn_vert+eng", config=custom_config)
                    
                    lines = [l.strip() for l in text.split('\n') if l.strip()]
                    if lines: all_regions_data.append([[l] for l in lines])
                    else: all_regions_data.append([[""]])
                except Exception: all_regions_data.append([["Error"]])
                    
            if options.get("extract_mode") == "text":
                merged_row = []
                for region_data in all_regions_data:
                    region_texts = []
                    for r in region_data:
                        for c in r:
                            val = str(c).strip()
                            if val and val != "Error": region_texts.append(val)
                    merged_row.append("\n".join(region_texts) if region_texts else "")
                merged_data = [merged_row] if merged_row else [[""]]
                final_data = [["ページ番号"] + ([f"範囲{idx+1}" for idx in range(len(cropped_images))] if crop_regions else ["抽出テキスト"])]
            else:
                merged_data = merge_2d_arrays_horizontally(all_regions_data)
                final_data = [["ページ番号"] + ([f"範囲{idx+1}" for idx in range(len(cropped_images))] if crop_regions else ["抽出テキスト"])]
            
            for row in merged_data: final_data.append([f"{page_num+1}/{total_pages}"] + row)
            
            save_path = os.path.join(save_dir, f"{base}_P{page_num+1}_OCR")
            if out_format == "xlsx":
                wb = Workbook(); ws = wb.active; ws.title = "OCR"
                for r_idx, r_data in enumerate(final_data, 1):
                    for c_idx, val in enumerate(r_data, 1): ws.cell(row=r_idx, column=c_idx, value=sanitize_excel_text(val))
                auto_adjust_excel_column_width(ws); wb.save(f"{save_path}.xlsx")
            elif out_format == "csv":
                with open(f"{save_path}.csv", "w", encoding="utf-8-sig", newline="") as f_out: csv.writer(f_out).writerows(final_data)
            elif out_format == "txt":
                with open(f"{save_path}.txt", "w", encoding="utf-8") as f_out:
                    for r in final_data: f_out.write("\t".join(r) + "\n")
        doc.close(); gc.collect()

# ==============================
# データ集約タスク (ローカル)
# ==============================
def aggregate_local_task(files, save_dir, options, ui):
    out_format = options.get("out_format", "xlsx")
    search_ext = "xlsx" if out_format in ["jpg", "png", "dxf", "svg", "tiff", "bmp"] else out_format
    run_timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if not files: raise Exception("対象が選択されていません。")
    search_exts = ["xlsx", "xlsm", "xls"] if search_ext == "xlsx" else [search_ext]
    target_files_set = set()
    for f in files:
        if os.path.isdir(f):
            for fn in os.listdir(f):
                if any(fn.lower().endswith(f".{ext}") for ext in search_exts) and "集約" not in fn and "結合" not in fn:
                    target_files_set.add(os.path.abspath(os.path.join(f, fn)))
        elif os.path.isfile(f):
            if any(f.lower().endswith(f".{ext}") for ext in search_exts): target_files_set.add(os.path.abspath(f))

    target_files = sorted(list(target_files_set))
    if not target_files: raise Exception("集約対象が見つかりません。")

    agg_header, agg_rows = ["元ファイル名"], []
    
    def map_to_master(fname, curr_header, curr_rows):
        safe_header = [str(h).strip() if h else f"列{i+1}" for i, h in enumerate(curr_header)]
        col_mapping = {}
        for i, h in enumerate(safe_header):
            if h not in agg_header: agg_header.append(h)
            col_mapping[i] = agg_header.index(h)
        for r in curr_rows:
            row = [""] * len(agg_header); row[0] = fname
            for i, val in enumerate(r):
                if i in col_mapping:
                    idx = col_mapping[i]
                    if idx >= len(row): row.extend([""] * (idx - len(row) + 1))
                    
                    if val is None or str(val).strip() == "None":
                        row[idx] = ""
                    else:
                        row[idx] = str(val).strip()
            agg_rows.append(row)

    for i, f in enumerate(target_files, 1):
        if ui.is_cancelled(): return
        ui.update_overall(i, len(target_files), f"集約中... ({i}/{len(target_files)})")
        ext = os.path.splitext(f)[1].lower().strip('.')
        fname = os.path.basename(f)
        try:
            if ext in ["xlsx", "xlsm"]:
                wb = openpyxl.load_workbook(f, data_only=True)
                for sheet in wb.sheetnames:
                    rows = [r for r in wb[sheet].iter_rows(values_only=True) if any(c is not None for c in r)]
                    if rows: map_to_master(fname, rows[0], rows[1:])
                wb.close()
            elif ext == "csv":
                with open(f, "r", encoding="utf-8-sig") as f_in: rows = list(csv.reader(f_in))
                if rows: map_to_master(fname, rows[0], rows[1:])
        except Exception: pass
        
    if len(agg_header) <= 1 or not agg_rows:
        raise Exception(
            f"集約するデータまたは見出し（ヘッダー）が見つからなかったため、ファイルは作成されませんでした。\n\n"
            f"【考えられる原因】\n"
            f"・選択したフォルダ/ファイルに「{search_ext.upper()}」形式の有効なデータが含まれていない\n"
            f"・対象ファイルの中身が完全に空である\n"
            f"・抽出処理に失敗し、表の列名（見出し）やデータが正しく生成されていない"
        )
        
    if save_dir:
        final_data = [agg_header] + [r + [""]*(len(agg_header)-len(r)) for r in agg_rows]
        apply_text_inheritance(final_data)
        
        if search_ext == "xlsx":
            wb = Workbook(); ws = wb.active; ws.title = "集約"
            for r_idx, r_data in enumerate(final_data, 1):
                for c_idx, val in enumerate(r_data, 1): ws.cell(row=r_idx, column=c_idx, value=sanitize_excel_text(val))
            auto_adjust_excel_column_width(ws); wb.save(os.path.join(save_dir, f"データ集約_{run_timestamp}.xlsx"))
        elif search_ext == "csv":
            with open(os.path.join(save_dir, f"データ集約_{run_timestamp}.csv"), "w", encoding="utf-8-sig", newline="") as f_out:
                csv.writer(f_out).writerows(final_data)

# ==============================
# データ単純結合タスク (ローカル)
# ==============================
def combine_local_task(files, save_dir, options, ui):
    out_format = options.get("out_format", "xlsx")
    search_ext = "xlsx" if out_format in ["jpg", "png", "dxf", "svg", "tiff", "bmp"] else out_format
    run_timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    if not files: raise Exception("対象が選択されていません。")
    search_exts = ["xlsx", "xlsm", "xls"] if search_ext == "xlsx" else [search_ext]
    target_files_set = set()
    for f in files:
        if os.path.isdir(f):
            for fn in os.listdir(f):
                if any(fn.lower().endswith(f".{ext}") for ext in search_exts) and "集約" not in fn and "結合" not in fn:
                    target_files_set.add(os.path.abspath(os.path.join(f, fn)))
        elif os.path.isfile(f):
            if any(f.lower().endswith(f".{ext}") for ext in search_exts): target_files_set.add(os.path.abspath(f))

    target_files = sorted(list(target_files_set))
    if not target_files: raise Exception("結合対象が見つかりません。")

    combined_rows = []

    for i, f in enumerate(target_files, 1):
        if ui.is_cancelled(): return
        ui.update_overall(i, len(target_files), f"結合中... ({i}/{len(target_files)})")
        ext = os.path.splitext(f)[1].lower().strip('.')
        fname = os.path.basename(f)
        try:
            if ext in ["xlsx", "xlsm"]:
                wb = openpyxl.load_workbook(f, data_only=True)
                for sheet in wb.sheetnames:
                    for r in wb[sheet].iter_rows(values_only=True):
                        # 空行以外は追加
                        if any(c is not None and str(c).strip() != "" for c in r):
                            combined_rows.append([fname] + [str(c) if c is not None else "" for c in r])
                wb.close()
            elif ext == "csv":
                with open(f, "r", encoding="utf-8-sig") as f_in:
                    for r in csv.reader(f_in):
                        if any(c.strip() != "" for c in r):
                            combined_rows.append([fname] + r)
        except Exception: pass
        
    if not combined_rows:
        raise Exception(
            f"結合するデータが見つからなかったため、ファイルは作成されませんでした。\n\n"
            f"【考えられる原因】\n"
            f"・選択したフォルダ/ファイルに「{search_ext.upper()}」形式の有効なデータが含まれていない\n"
            f"・対象ファイルの中身が完全に空である"
        )
        
    if save_dir:
        if search_ext == "xlsx":
            wb = Workbook(); ws = wb.active; ws.title = "単純結合"
            for r_idx, r_data in enumerate(combined_rows, 1):
                for c_idx, val in enumerate(r_data, 1): ws.cell(row=r_idx, column=c_idx, value=sanitize_excel_text(val))
            auto_adjust_excel_column_width(ws); wb.save(os.path.join(save_dir, f"データ結合_{run_timestamp}.xlsx"))
        elif search_ext == "csv":
            with open(os.path.join(save_dir, f"データ結合_{run_timestamp}.csv"), "w", encoding="utf-8-sig", newline="") as f_out:
                csv.writer(f_out).writerows(combined_rows)