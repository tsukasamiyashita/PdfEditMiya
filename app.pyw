# -*- coding: utf-8 -*-
import os, sys, threading, io, gc, cv2, csv, time, json, ast, difflib, re, random
import numpy as np
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, Menu
import tkinter.scrolledtext as st
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
import openpyxl  # エラー修正: aggregate_only_task 用にインポート
from openpyxl import Workbook
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import fitz  # PyMuPDF
import ezdxf
from PIL import Image, ImageTk
import pytesseract
import google.generativeai as genai

# ==============================
# 基本設定 & カラーパレット
# ==============================
APP_TITLE, VERSION = "PdfEditMiya", "v2.0.0"
WINDOW_WIDTH, WINDOW_HEIGHT = 700, 820

BG_COLOR, CARD_BG = "#F0F4F8", "#FFFFFF"
PRIMARY, PRIMARY_HOVER = "#0D6EFD", "#0B5ED7"
TEXT_COLOR, MUTED_TEXT, BORDER_COLOR = "#212529", "#6C757D", "#DEE2E6"

# エラー修正: 不足していたカラー変数をすべて追加
SUCCESS, ERROR = "#198754", "#DC3545"
COLOR_SUCCESS, COLOR_SUCCESS_HOVER = "#198754", "#157347"
COLOR_INFO, COLOR_INFO_HOVER = "#0DCAF0", "#0BACCE"
COLOR_WARNING, COLOR_WARNING_HOVER = "#FFC107", "#E0A800"
COLOR_DANGER, COLOR_DANGER_HOVER = "#DC3545", "#B02A37"
COLOR_PURPLE, COLOR_PURPLE_HOVER = "#6F42C1", "#59339D"

USER_HOME = os.path.expanduser("~")
API_KEY_FILE = os.path.join(USER_HOME, ".pdfeditmiya_api_key.txt")
AI_HELP_TEXT = "【 AI抽出機能の使い方と準備 】\nPDF内の表データや手書き文字を解析し、Excel(xlsx)・CSV・テキストデータとして抽出する機能です。\n用途に合わせてGemini APIかTesseractをご利用ください。"

# ==============================
# グローバル状態
# ==============================
selected_files, selected_folder, current_mode = [], "", None
preset_save_dir, selected_crop_regions = "", []
processing_popup, overall_label, overall_progress = None, None, None
file_label, file_progress, cancelled = None, None, False

# ==============================
# UI コントローラー & ヘルパー関数
# ==============================
class UIController:
    def update_overall(self, step, max_val=None, text=None):
        def _task():
            if processing_popup and processing_popup.winfo_exists():
                if max_val is not None: overall_progress["maximum"] = max_val
                overall_progress["value"] = step
                if text: overall_label.config(text=text)
        root.after(0, _task)
    def set_indeterminate(self, text=None):
        def _task():
            if processing_popup and processing_popup.winfo_exists():
                file_progress.config(mode="indeterminate"); file_progress.start(15)
                if text: file_label.config(text=text)
        root.after(0, _task)
    def set_determinate(self, step, max_val=None, text=None):
        def _task():
            if processing_popup and processing_popup.winfo_exists():
                file_progress.stop(); file_progress.config(mode="determinate")
                if max_val is not None: file_progress["maximum"] = max_val
                file_progress["value"] = step
                if text: file_label.config(text=text)
        root.after(0, _task)

def resource_path(relative_path):
    try: base_path = sys._MEIPASS
    except Exception: base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

def auto_adjust_excel_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    for line in str(cell.value).split('\n'):
                        length = sum(2 if ord(c) > 255 else 1 for c in line)
                        if length > max_length: max_length = length
            except: pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 60)

def analyze_column_profile(col_data):
    if not col_data: return {"pure_num_ratio": 0.0, "fraction_ratio": 0.0, "avg_num_len": 0.0, "is_text": True}
    pure_num_cnt, fraction_cnt, total_num_len, total_cells = 0, 0, 0, 0
    for val in col_data:
        s = str(val).strip()
        if not s or s == "None": continue
        total_cells += 1
        if re.match(r'^\d+/\d+$', s):
            fraction_cnt += 1; total_num_len += len(s)
        else:
            s_clean = s.replace(",", "").replace(".", "", 1).replace("-", "", 1)
            if s_clean.isdigit(): pure_num_cnt += 1; total_num_len += len(s_clean)
    if total_cells == 0: return {"pure_num_ratio": 0.0, "fraction_ratio": 0.0, "avg_num_len": 0.0, "is_text": True}
    return {
        "pure_num_ratio": pure_num_cnt / total_cells,
        "fraction_ratio": fraction_cnt / total_cells,
        "avg_num_len": total_num_len / (pure_num_cnt + fraction_cnt) if (pure_num_cnt + fraction_cnt) > 0 else 0.0,
        "is_text": ((pure_num_cnt / total_cells) < 0.2 and (fraction_cnt / total_cells) < 0.2)
    }

def get_profile_similarity(p1, p2):
    diff_pure = abs(p1["pure_num_ratio"] - p2["pure_num_ratio"])
    diff_frac = abs(p1["fraction_ratio"] - p2["fraction_ratio"])
    max_len = max(p1["avg_num_len"], p2["avg_num_len"])
    diff_len = abs(p1["avg_num_len"] - p2["avg_num_len"]) / max_len if max_len > 0 else 0.0
    return max(0.0, 1.0 - (diff_pure * 0.4 + diff_frac * 0.4 + diff_len * 0.2))

def parse_row_data(row_data):
    if isinstance(row_data, list) and len(row_data) == 1: row_data = row_data[0]
    if isinstance(row_data, str):
        row_data = row_data.strip()
        if (row_data.startswith('(') and row_data.endswith(')')) or (row_data.startswith('[') and row_data.endswith(']')):
            try:
                parsed = ast.literal_eval(row_data)
                if isinstance(parsed, (list, tuple)): return [str(x) if x is not None else "" for x in parsed]
            except:
                return [x.strip().strip("'\"") for x in row_data.strip("()[]").split(",")]
        return [row_data]
    if isinstance(row_data, tuple): return [str(x) if x is not None else "" for x in row_data]
    if not isinstance(row_data, list): return [str(row_data)]
    return [str(x) if x is not None else "" for x in row_data]

def apply_text_inheritance(final_aggregated_data):
    if len(final_aggregated_data) <= 1: return
    def is_text_to_inherit(text):
        s = str(text).strip()
        if not s or s in ["〃", "”", "\"", "''", "””", "''", "同上"]: return False
        return bool(re.search(r'[a-zA-Zａ-ｚＡ-Ｚぁ-んァ-ン一-龥]', s))
    header = final_aggregated_data[0]
    skip_cols = {idx for idx, h in enumerate(header) if "備考" in str(h)}
    for col_idx in range(1, len(header)):
        if col_idx in skip_cols: continue
        last_text = ""
        for row_idx in range(1, len(final_aggregated_data)):
            cell_val = str(final_aggregated_data[row_idx][col_idx]).strip()
            if cell_val in ["", "None", "〃", "”", "\"", "''", "””", "''", "同上"]:
                if last_text: final_aggregated_data[row_idx][col_idx] = last_text
            else:
                last_text = cell_val if is_text_to_inherit(cell_val) else ""

def merge_2d_arrays_horizontally(arrays_list):
    if not arrays_list: return []
    max_rows = max((len(arr) for arr in arrays_list), default=0)
    merged = []
    region_max_cols = [max((len(row) for row in arr), default=0) if arr else 0 for arr in arrays_list]
    for r in range(max_rows):
        merged_row = []
        for i, arr in enumerate(arrays_list):
            max_c = region_max_cols[i]
            if arr and r < len(arr):
                row_data = list(arr[r])
                row_data += [""] * (max_c - len(row_data))
                merged_row.extend(row_data)
            else: merged_row.extend([""] * max_c)
        merged.append(merged_row)
    return merged
# ==============================
# コア処理関数群 (PDF編集・標準抽出)
# ==============================
def merge_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    writer = PdfWriter()
    ui.update_overall(0, len(files), f"全体の進捗 ( 0 / {len(files)} ファイル )")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ファイルを結合中... ( {j} / {len(reader.pages)} ページ )")
            writer.add_page(p)
    ui.set_determinate(1, 1, "PDFを保存中...")
    if save_dir:
        with open(os.path.join(save_dir, f"{options.get('folder_name', 'Merged')}_Merge.pdf"), "wb") as out: writer.write(out)

def split_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(reader.pages))))
        for n, p in enumerate(reader.pages, 1):
            ui.set_determinate(n, len(reader.pages), f"ファイルを分割中... ( {n} / {len(reader.pages)} ページ )")
            writer = PdfWriter(); writer.add_page(p)
            with open(os.path.join(save_dir, f"{base}_Split_{str(n).zfill(digits)}.pdf"), "wb") as out: writer.write(out)

def rotate_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f); writer = PdfWriter()
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ページを回転中... ( {j} / {len(reader.pages)} ページ )")
            p.rotate(options.get("rotate_deg", 270)); writer.add_page(p)
        base = os.path.splitext(os.path.basename(f))[0]
        with open(os.path.join(save_dir, f"{base}_Rotate.pdf"), "wb") as out: writer.write(out)

def extract_text_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(f))[0]; text_list = []
        with pdfplumber.open(f) as pdf:
            for j, p in enumerate(pdf.pages, 1):
                ui.set_determinate(j, len(pdf.pages), f"テキストを抽出中... ( {j} / {len(pdf.pages)} ページ )")
                if crop_regions:
                    page_texts = []
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * p.width, min(ry1, ry2) * p.height, max(rx1, rx2) * p.width, max(ry1, ry2) * p.height
                        txt = p.crop((x0, top, x1, bottom)).extract_text()
                        if txt: page_texts.append(txt)
                    if page_texts: text_list.append("\n".join(page_texts))
                else:
                    txt = p.extract_text()
                    if txt: text_list.append(txt)
        with open(os.path.join(save_dir, f"{base}_Text.txt"), "w", encoding="utf-8") as out: out.write("\n".join(text_list))

def convert_to_excel_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    border_style = Side(border_style="thin", color="000000")
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        wb = Workbook(); wb.remove(wb.active)
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"表データをExcelへ変換中... ( {page_idx} / {len(pdf.pages)} ページ )")
                tables = []
                if crop_regions:
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * page.width, min(ry1, ry2) * page.height, max(rx1, rx2) * page.width, max(ry1, ry2) * page.height
                        tbls = page.crop((x0, top, x1, bottom)).extract_tables()
                        if tbls: tables.extend(tbls)
                else: tables = page.extract_tables()
                if not tables: continue
                ws = wb.create_sheet(f"Page_{str(page_idx).zfill(digits)}"); current_row = 1
                for table in tables:
                    for row_data in table:
                        for col_idx, cell_value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_idx, value=str(cell_value).strip() if cell_value else "")
                            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                        current_row += 1
                    current_row += 2
                auto_adjust_excel_column_width(ws)
        if len(wb.sheetnames) > 0: wb.save(os.path.join(save_dir, f"{os.path.splitext(os.path.basename(pdf_path))[0]}_Excel.xlsx"))

def convert_to_csv_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"表データをCSVへ変換中... ( {page_idx} / {len(pdf.pages)} ページ )")
                tables = []
                if crop_regions:
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * page.width, min(ry1, ry2) * page.height, max(rx1, rx2) * page.width, max(ry1, ry2) * page.height
                        tbls = page.crop((x0, top, x1, bottom)).extract_tables()
                        if tbls: tables.extend(tbls)
                else: tables = page.extract_tables()
                if not tables: continue
                with open(os.path.join(save_dir, f"{base}_Page_{str(page_idx).zfill(digits)}_CSV.csv"), "w", encoding="utf-8-sig", newline="") as f_out:
                    writer = csv.writer(f_out)
                    for table in tables:
                        for row_data in table: writer.writerow([str(cell).strip() if cell else "" for cell in row_data])
                        writer.writerow([]) 

def convert_to_image_jpg(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "jpg")
def convert_to_image_png(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "png")
def _convert_image(files, save_dir, options, ui, ext):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        doc = fitz.open(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(doc))))
        for n, page in enumerate(doc, 1):
            ui.set_determinate(n, len(doc), f"画像へ変換中... ( {n} / {len(doc)} ページ )")
            n_str = str(n).zfill(digits)
            pix = page.get_pixmap(dpi=200)
            if crop_regions:
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
                elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)
                h, w = img_array.shape[:2]
                for idx, (rx1, ry1, rx2, ry2) in enumerate(crop_regions):
                    x1, y1 = int(min(rx1, rx2) * w), int(min(ry1, ry2) * h)
                    x2, y2 = int(max(rx1, rx2) * w), int(max(ry1, ry2) * h)
                    Image.fromarray(img_array[y1:y2, x1:x2]).save(os.path.join(save_dir, f"{base}_{n_str}_crop{idx+1}.{ext}"))
            else: pix.save(os.path.join(save_dir, f"{base}_{n_str}.{ext}"))

def convert_to_dxf(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    def in_crop(x, y, w, h):
        if not crop_regions: return True
        for (rx1, ry1, rx2, ry2) in crop_regions:
            if min(rx1, rx2)*w <= x <= max(rx1, rx2)*w and min(ry1, ry2)*h <= y <= max(ry1, ry2)*h: return True
        return False

    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        try:
            doc = fitz.open(f); dwg = ezdxf.new('R2010'); msp = dwg.modelspace()
            for page_num, page in enumerate(doc, 1):
                ui.set_determinate(page_num, len(doc), f"DXFへ変換中... ( {page_num} / {len(doc)} ページ )")
                h, w = page.rect.height, page.rect.width
                paths = page.get_drawings()
                is_vector_rich = len(paths) > 0
                if is_vector_rich:
                    for path in paths:
                        for item in path["items"]:
                            if item[0] == "l":
                                p1, p2 = item[1], item[2]
                                if in_crop(p1.x, p1.y, w, h) or in_crop(p2.x, p2.y, w, h): msp.add_line((p1.x, h - p1.y), (p2.x, h - p2.y))
                            elif item[0] == "re":
                                rect = item[1]
                                if in_crop(rect.x0, rect.y0, w, h) or in_crop(rect.x1, rect.y1, w, h):
                                    msp.add_lwpolyline([(rect.x0, h - rect.y0), (rect.x1, h - rect.y0), (rect.x1, h - rect.y1), (rect.x0, h - rect.y1)], close=True)
                            elif item[0] == "c":
                                p1, p2, p3, p4 = item[1], item[2], item[3], item[4]
                                if any(in_crop(pt.x, pt.y, w, h) for pt in [p1, p2, p3, p4]):
                                    msp.add_spline([(p1.x, h - p1.y), (p2.x, h - p2.y), (p3.x, h - p3.y), (p4.x, h - p4.y)])
                if not is_vector_rich or len(paths) < 5:
                    pix = page.get_pixmap(dpi=300)
                    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY) if pix.n >= 3 else img
                    if crop_regions:
                        mask = np.zeros_like(gray)
                        for (rx1, ry1, rx2, ry2) in crop_regions:
                            mask[int(min(ry1, ry2)*gray.shape[0]):int(max(ry1, ry2)*gray.shape[0]), int(min(rx1, rx2)*gray.shape[1]):int(max(rx1, rx2)*gray.shape[1])] = 255
                        gray = cv2.bitwise_and(gray, mask)
                    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
                    binary = cv2.morphologyEx(cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2), cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
                    contours, _ = cv2.findContours(binary, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
                    scale_x, scale_y = w / pix.w, h / pix.h
                    for cnt in contours:
                        if cv2.contourArea(cnt) < 15: continue 
                        pts = [(p[0][0] * scale_x, h - p[0][1] * scale_y) for p in cv2.approxPolyDP(cnt, 0.003 * cv2.arcLength(cnt, True), True)]
                        if len(pts) > 1: msp.add_lwpolyline(pts, close=True)
            ui.set_determinate(len(doc), len(doc), "DXFファイルを保存中...")
            dwg.saveas(os.path.join(save_dir, f"{os.path.splitext(os.path.basename(f))[0]}_CAD.dxf"))
        except Exception as e: print(f"DXF Conversion Error: {e}")
# ==============================
# AI抽出・データ集約タスク
# ==============================
def check_tesseract_installation():
    if sys.platform.startswith("win"):
        tess_path = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
        if os.path.exists(tess_path): pytesseract.pytesseract.tesseract_cmd = tess_path
    try: pytesseract.get_tesseract_version()
    except Exception: raise Exception("Tesseract OCRがインストールされていないか、PATHが通っていません。")

def extract_tesseract_task(files, save_dir, options, ui):
    check_tesseract_installation()
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    out_format = options.get("out_format", "xlsx")
    crop_regions = options.get("crop_regions", [])
    
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(f))[0]
        doc = fitz.open(f)
        digits = max(2, len(str(len(doc))))
        for page_num in range(len(doc)):
            ui.set_indeterminate(f"Tesseractで解析中... ( {page_num+1} / {len(doc)} ページ )")
            pix = doc[page_num].get_pixmap(dpi=300)
            img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
            if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
            elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)
            
            cropped_images = []
            if crop_regions:
                h, w = img_array.shape[:2]
                for (rx1, ry1, rx2, ry2) in crop_regions:
                    cropped_images.append(img_array[int(min(ry1, ry2)*h):int(max(ry1, ry2)*h), int(min(rx1, rx2)*w):int(max(rx1, rx2)*w)])
            else: cropped_images.append(img_array)
                
            all_regions_data = []
            for crop_img in cropped_images:
                try:
                    text = pytesseract.image_to_string(Image.fromarray(crop_img), lang="jpn+eng")
                    lines = [line.strip() for line in text.split('\n') if line.strip()]
                    if lines:
                        if (crop_regions and crop_img.shape[0] > crop_img.shape[1] * 1.5) or (len(lines) > 1 and all(len(l) <= 2 for l in lines)):
                            all_regions_data.append([["".join(lines)]])
                        else: all_regions_data.append([[l] for l in lines])
                    else: all_regions_data.append([[""]])
                except Exception as e: raise Exception(f"Tesseract OCRエラー: {e}")
                    
            merged_data = merge_2d_arrays_horizontally(all_regions_data)
            final_data = []
            page_info = f"{page_num+1}/{len(doc)}"
            if crop_regions:
                final_data.append(["ページ番号"] + [f"抽出範囲{idx+1}" for idx in range(len(cropped_images))])
            else:
                final_data.append(["ページ番号", "抽出テキスト"])
            for row in merged_data: final_data.append([page_info] + row)
            
            save_path = os.path.join(save_dir, f"{base}_Page_{str(page_num+1).zfill(digits)}_Tesseract抽出")
            if out_format == "xlsx":
                wb = Workbook(); ws = wb.active; ws.title = f"Page_{str(page_num+1).zfill(digits)}"
                for r_idx, row_data in enumerate(final_data, 1):
                    for c_idx, val in enumerate(row_data, 1): ws.cell(row=r_idx, column=c_idx, value=str(val))
                auto_adjust_excel_column_width(ws); wb.save(f"{save_path}.xlsx")
            elif out_format == "csv":
                with open(f"{save_path}.csv", "w", encoding="utf-8-sig", newline="") as f_out: csv.writer(f_out).writerows(final_data)
            elif out_format == "txt":
                with open(f"{save_path}.txt", "w", encoding="utf-8") as f_out:
                    for row_data in final_data: f_out.write("\t".join(row_data) + "\n")
        doc.close(); gc.collect()
        ui.set_determinate(len(doc), len(doc), "完了")

def extract_gemini_task(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    genai.configure(api_key=options.get("api_key", ""))
    models_to_try = options.get("models_to_try", [])
    out_format = options.get("out_format", "xlsx")
    crop_regions = options.get("crop_regions", [])

    if out_format in ["csv", "xlsx"]:
        if crop_regions:
            prompt = """あなたは優秀なデータ入力オペレーターです。添付画像のテキストを読み取りJSONを作成してください。
            【特別ルール】出力は「1つの列」にまとめ、セル分けしないでください。縦書きテキストは繋げて横書きに変換してください。
            【出力形式】 {"rows": ["1行目のテキスト...", "2行目のテキスト..."]}"""
        else:
            prompt = """あなたは優秀なデータ入力オペレーターです。表画像を読み取りJSONを作成してください。
            【特別ルール】見た目より「物理的な列の区切り(罫線・空白)」を最優先し分割してください。空白セルは `""` を挿入してください。縦書きテキストは繋げて横書きにしてください。
            【出力形式】 {"header": ["列1", ...], "rows": [["データ1", ""], ...]}"""
        generation_config = {"response_mime_type": "application/json"}
    else:
        prompt = "この画像の文字を可能な限り正確に読み取りプレーンテキストとして出力してください。縦書きの文章は横書きに変換してください。"
        generation_config = None

    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(f))[0]
        doc = fitz.open(f)
        digits = max(2, len(str(len(doc))))
        for page_num in range(len(doc)):
            pix = doc[page_num].get_pixmap(dpi=300)
            img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
            if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
            elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)

            cropped_images = []
            if crop_regions:
                h, w = img_array.shape[:2]
                for (rx1, ry1, rx2, ry2) in crop_regions:
                    cropped_images.append(img_array[int(min(ry1, ry2)*h):int(max(ry1, ry2)*h), int(min(rx1, rx2)*w):int(max(rx1, rx2)*w)])
            else: cropped_images.append(img_array)
                
            all_regions_data = []
            for region_idx, crop_img_array in enumerate(cropped_images):
                gray = cv2.cvtColor(crop_img_array, cv2.COLOR_RGB2GRAY)
                blurred = cv2.GaussianBlur(gray, (3, 3), 0)
                clean_bg = np.where(cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 15, 5) == 255, 255, gray)
                enhanced = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8, 8)).apply(clean_bg.astype(np.uint8))
                img = Image.fromarray(cv2.cvtColor(cv2.addWeighted(enhanced, 1.5, cv2.GaussianBlur(enhanced, (0, 0), 2), -0.5, 0), cv2.COLOR_GRAY2RGB))
                
                extracted_text, success, last_error = "", False, ""
                for attempt in range(4):
                    for model_name in models_to_try:
                        ui.set_indeterminate(f"AI解析中... ( {page_num+1}/{len(doc)}頁 | 領域 {region_idx+1}/{len(cropped_images)} )")
                        try:
                            model = genai.GenerativeModel(model_name)
                            response = model.generate_content([prompt, img], generation_config=generation_config) if generation_config else model.generate_content([prompt, img])
                            if not response.parts: raise Exception("安全フィルタブロック")
                            extracted_text = response.text.strip(); success = True; break
                        except Exception as api_err: last_error = str(api_err); continue
                    if success: break
                    time.sleep((2 ** attempt) + random.uniform(0.5, 1.5))

                if success:
                    time.sleep(4.0) 
                    if out_format in ["xlsx", "csv"]:
                        try:
                            data = json.loads(extracted_text)
                            if crop_regions:
                                rows = data.get("rows", []) if isinstance(data, dict) else data
                                page_data_to_write = []
                                h_c, w_c = crop_img_array.shape[:2]
                                clean_rows = ["".join([l.strip() for l in (" ".join(str(x) for x in r) if isinstance(r, list) else str(r)).split('\n') if l.strip()]) if all(len(l)<=2 for l in (" ".join(str(x) for x in r) if isinstance(r, list) else str(r)).split('\n')) else (" ".join(str(x) for x in r) if isinstance(r, list) else str(r)) for r in rows]
                                if h_c > w_c * 1.5 and all(len(x.strip()) <= 2 for x in clean_rows if x.strip()): page_data_to_write.append(["".join(clean_rows)])
                                else:
                                    for val in clean_rows: page_data_to_write.append([val])
                                all_regions_data.append(page_data_to_write)
                            else:
                                header = data.get("header", [])
                                rows = data.get("rows", [])
                                if not header and not rows and isinstance(data, list) and data:
                                    header, rows = (data[0] if isinstance(data[0], list) else [str(data[0])]), data[1:]
                                safe_header = [str(x).strip() for x in header] if isinstance(header, list) else []
                                page_col_count = max(len(safe_header), max([len(r) for r in rows if isinstance(r, list)] + [0]))
                                if not safe_header: safe_header = [f"列{i+1}" for i in range(page_col_count)]
                                page_data_to_write = [(safe_header + [""] * page_col_count)[:page_col_count]]
                                for row_data in rows:
                                    safe_row = [("".join([l.strip() for l in str(v).split('\n')]) if '\n' in str(v) and len([l for l in str(v).split('\n') if l.strip()]) > 1 and all(len(l) <= 2 for l in str(v).split('\n') if l.strip()) else str(v)) for v in (parse_row_data(row_data) + [""] * page_col_count)[:page_col_count]]
                                    if any(v != "" for v in safe_row): page_data_to_write.append(safe_row)
                                all_regions_data.append(page_data_to_write)
                        except: all_regions_data.append([[f"JSONパースエラー"]])
                    else: all_regions_data.append([[line] for line in extracted_text.split('\n')])
                else: all_regions_data.append([[f"抽出失敗: {last_error}"]])
            
            merged_data = merge_2d_arrays_horizontally(all_regions_data)
            final_data = []
            page_info = f"{page_num+1}/{len(doc)}"
            if crop_regions:
                final_data.append(["ページ番号"] + [f"抽出範囲{idx+1}" for idx in range(len(cropped_images))])
                for row in merged_data: final_data.append([page_info] + row)
            else:
                for r_idx, row in enumerate(merged_data):
                    final_data.append(["ページ番号" if r_idx == 0 and out_format in ["xlsx", "csv"] else page_info] + row)
            
            save_path = os.path.join(save_dir, f"{base}_Page_{str(page_num+1).zfill(digits)}_AI抽出")
            if out_format == "xlsx":
                wb = Workbook(); ws = wb.active; ws.title = f"Page_{str(page_num+1).zfill(digits)}"
                for r_idx, row_data in enumerate(final_data, 1):
                    for c_idx, val in enumerate(row_data, 1): ws.cell(row=r_idx, column=c_idx, value=str(val))
                auto_adjust_excel_column_width(ws); wb.save(f"{save_path}.xlsx")
            elif out_format == "csv":
                with open(f"{save_path}.csv", "w", encoding="utf-8-sig", newline="") as f_out: csv.writer(f_out).writerows(final_data)
            elif out_format == "txt":
                with open(f"{save_path}.txt", "w", encoding="utf-8") as f_out:
                    for row_data in final_data: f_out.write("\t".join(row_data) + "\n")
        doc.close(); gc.collect()
        ui.set_determinate(len(doc), len(doc), "完了")

def aggregate_only_task(files, save_dir, options, ui):
    out_format = options.get("out_format", "xlsx")
    search_ext = "xlsx" if out_format in ["jpg", "png", "dxf"] else out_format
    target_files = []
    for f in files:
        ext = os.path.splitext(f)[1].lower()
        if ext == ".pdf":
            base = os.path.splitext(f)[0]
            dir_path = os.path.dirname(f)
            target_files.extend([os.path.join(dir_path, fn) for fn in os.listdir(dir_path) if fn.startswith(base) and fn.endswith(f".{search_ext}") and any(x in fn for x in ["_AI抽出", "_Tesseract抽出", "_Excel", "_CSV", "_Text"])])
        elif ext == f".{search_ext}": target_files.append(f)
    target_files = sorted(list(set(target_files)), key=lambda x: os.path.getmtime(x))
    if not target_files: raise Exception(f"指定された出力形式 (.{search_ext}) のデータが見つかりません。")

    agg_header, agg_rows, master_profiles, agg_texts = ["元ファイル名"], [], {}, []
    def map_to_master(fname, curr_header, curr_rows):
        safe_header = [str(h).strip() if h is not None else "" or f"列{i+1}" for i, h in enumerate(parse_row_data(curr_header))]
        col_count = len(safe_header)
        col_data_list = [[] for _ in range(col_count)]
        parsed_rows = []
        for r in curr_rows:
            r_list = parse_row_data(r)
            parsed_rows.append(r_list)
            for i, val in enumerate(r_list):
                if i < col_count: col_data_list[i].append(val)
        curr_profiles = [analyze_column_profile(col_data_list[i]) for i in range(col_count)]
        col_mapping, mapped_master_indices = {}, set()
        for i, h in enumerate(safe_header):
            best_idx, best_score = -1, -1
            for m_idx, m_h in enumerate(agg_header):
                if m_idx == 0 or m_idx in mapped_master_indices: continue
                if m_idx in master_profiles:
                    p_c, p_m = curr_profiles[i], master_profiles[m_idx]
                    if p_c["is_text"] and p_m["is_text"]: score = 1.0 if i == (m_idx - 1) else 0.5 - abs(i - (m_idx - 1)) * 0.1
                    else: score = get_profile_similarity(p_c, p_m) - abs(i - (m_idx - 1)) * 0.05
                    if score > best_score and score > 0.4: best_idx, best_score = m_idx, score
            if best_idx != -1:
                col_mapping[i] = best_idx; mapped_master_indices.add(best_idx)
                old_p, new_p = master_profiles[best_idx], curr_profiles[i]
                master_profiles[best_idx] = {"pure_num_ratio": (old_p["pure_num_ratio"]+new_p["pure_num_ratio"])/2, "fraction_ratio": (old_p["fraction_ratio"]+new_p["fraction_ratio"])/2, "avg_num_len": (old_p["avg_num_len"]+new_p["avg_num_len"])/2, "is_text": old_p["is_text"] and new_p["is_text"]}
            else:
                agg_header.append(h); new_idx = len(agg_header) - 1
                col_mapping[i] = new_idx; mapped_master_indices.add(new_idx)
                master_profiles[new_idx] = curr_profiles[i]
        for r in parsed_rows:
            row = [""] * len(agg_header)
            row[0] = fname
            for i, val in enumerate(r):
                if i in col_mapping:
                    m_idx = col_mapping[i]
                    if m_idx >= len(row): row.extend([""] * (m_idx - len(row) + 1))
                    row[m_idx] = str(val).strip() if val is not None and str(val).strip() != "None" else ""
            if any(v != "" for v in row[1:]): agg_rows.append(row)

    for i, f in enumerate(target_files, 1):
        ui.update_overall(i, len(target_files), f"データを集約中... ( {i} / {len(target_files)} ファイル )")
        ui.set_determinate(50, 100, f"読み込み中: {os.path.basename(f)}")
        fname = re.sub(r'(_Page_\d+)?(_AI抽出|_Tesseract抽出|_Excel|_CSV|_Text)\.' + search_ext + '$', '.pdf', os.path.basename(f))
        try:
            if search_ext == "xlsx":
                wb = openpyxl.load_workbook(f, data_only=True)
                for sheet in wb.sheetnames:
                    rows = list(wb[sheet].iter_rows(values_only=True))
                    if rows: map_to_master(fname, rows[0], rows[1:])
                wb.close()
            elif search_ext == "csv":
                with open(f, "r", encoding="utf-8-sig") as f_in:
                    rows = list(csv.reader(f_in))
                    if rows: map_to_master(fname, rows[0], rows[1:])
            elif search_ext == "txt":
                with open(f, "r", encoding="utf-8") as f_in: agg_texts.append(f"[{fname}]\n{f_in.read()}")
        except Exception as e: print(f"Read Error: {e}")
        ui.set_determinate(100, 100, "完了"); time.sleep(0.05)
        
    if len(target_files) > 0 and save_dir:
        ui.set_indeterminate("集約データを保存中...")
        if search_ext in ["xlsx", "csv"]:
            final_data = [agg_header] + [(r + [""] * len(agg_header))[:len(agg_header)] for r in agg_rows]
            apply_text_inheritance(final_data)
            if search_ext == "xlsx" and len(final_data) > 1:
                wb = Workbook(); ws = wb.active; ws.title = "集約データ"
                for r_idx, r_data in enumerate(final_data, 1):
                    for c_idx, val in enumerate(r_data, 1): ws.cell(row=r_idx, column=c_idx, value=str(val).strip())
                auto_adjust_excel_column_width(ws); wb.save(os.path.join(save_dir, "データ集約.xlsx"))
            elif search_ext == "csv" and len(final_data) > 1:
                with open(os.path.join(save_dir, "データ集約.csv"), "w", encoding="utf-8-sig", newline="") as f_out: csv.writer(f_out).writerows(final_data)
        elif search_ext == "txt" and agg_texts:
            with open(os.path.join(save_dir, "データ集約.txt"), "w", encoding="utf-8") as f_out: f_out.write("\n\n".join(agg_texts))
# ==============================
# UI制御＆イベント・画面構築
# ==============================
def get_api_key():
    if os.path.exists(API_KEY_FILE):
        with open(API_KEY_FILE, "r", encoding="utf-8") as f: return f.read().strip()
    return None

def test_api_key_ui():
    key = api_key_var.get().strip()
    if not key: return messagebox.showwarning("警告", "APIキーが入力されていません。")
    genai.configure(api_key=key)
    try:
        genai.GenerativeModel("gemini-1.5-flash").generate_content("Test")
        with open(API_KEY_FILE, "w", encoding="utf-8") as f: f.write(key)
        messagebox.showinfo("認証成功！", "APIキーは正しく認識されました。")
    except Exception as e: messagebox.showerror("通信エラー", f"APIキー確認中にエラーが発生しました。\n{e}")

def show_message(msg, color=PRIMARY):
    def _task():
        win = tk.Toplevel(root)
        win.geometry("260x90"); win.configure(bg=CARD_BG); win.attributes("-topmost", True)
        x = root.winfo_x() + (WINDOW_WIDTH // 2) - 130; y = root.winfo_y() + (WINDOW_HEIGHT // 2) - 45
        win.geometry(f"+{x}+{y}"); win.overrideredirect(True)
        frame = tk.Frame(win, bg=CARD_BG, highlightbackground=color, highlightthickness=2); frame.pack(expand=True, fill=tk.BOTH)
        ttk.Label(frame, text=msg, foreground=color, font=("Segoe UI", 10, "bold"), background=CARD_BG, wraplength=240).pack(expand=True)
        win.after(2500, win.destroy)
    root.after(0, _task)

def show_processing(total_files=1):
    global processing_popup, overall_label, overall_progress, file_label, file_progress
    processing_popup = tk.Toplevel(root)
    processing_popup.title("処理を実行中..."); processing_popup.geometry("440x210"); processing_popup.configure(bg=CARD_BG); processing_popup.grab_set()
    x = root.winfo_x() + (WINDOW_WIDTH // 2) - 220; y = root.winfo_y() + (WINDOW_HEIGHT // 2) - 105
    processing_popup.geometry(f"+{x}+{y}")
    overall_label = ttk.Label(processing_popup, text=f"全体の進捗 ( 0 / {total_files} ファイル )", font=("Segoe UI", 10, "bold"), background=CARD_BG, foreground=PRIMARY)
    overall_label.pack(pady=(25, 5))
    overall_progress = ttk.Progressbar(processing_popup, mode="determinate", maximum=total_files, length=380); overall_progress.pack(pady=(0, 20))
    file_label = ttk.Label(processing_popup, text="現在のファイルを準備中...", font=("Segoe UI", 9), background=CARD_BG, foreground=MUTED_TEXT); file_label.pack(pady=(5, 5))
    file_progress = ttk.Progressbar(processing_popup, mode="determinate", maximum=1, length=380); file_progress.pack(pady=(0, 10))

def close_processing():
    def _task():
        global processing_popup
        if processing_popup: processing_popup.destroy(); processing_popup = None
    root.after(0, _task)

def run_task(func):
    global cancelled; cancelled = False
    try:
        files = selected_files if current_mode == "file" else ([os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith((".pdf", ".xlsx", ".csv", ".txt"))] if selected_folder else [])
        if not files: return
        save_dir = os.path.dirname(files[0]) if save_option.get() == 1 else preset_save_dir
        options = {
            "rotate_deg": rotate_option.get(), "crop_regions": selected_crop_regions, "out_format": output_format_var.get(),
            "folder_name": os.path.basename(selected_folder) if selected_folder else "Merged",
            "api_key": api_key_var.get().strip(), "models_to_try": ['gemini-1.5-pro', 'gemini-1.5-flash'] if engine_var.get() == "Gemini" else []
        }
        func(files, save_dir, options, UIController())
        close_processing()
        if not cancelled: show_message("✅ 処理が完了しました", SUCCESS)
    except Exception as e:
        print(f"Error: {e}"); close_processing(); show_message(f"❌ エラーが発生しました\n{str(e)[:40]}...", ERROR)

def safe_run(func):
    files = selected_files if current_mode == "file" else ([os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith((".pdf", ".xlsx", ".csv", ".txt"))] if selected_folder else [])
    if not files: return
    global preset_save_dir
    if save_option.get() == 2 and not preset_save_dir:
        folder = filedialog.askdirectory(title="保存先フォルダを選択")
        if not folder: return
        preset_save_dir = folder; save_label.config(text=preset_save_dir)
    show_processing(len(files))
    threading.Thread(target=run_task, args=(func,), daemon=True).start()

def run_selected_extraction():
    engine = engine_var.get(); fmt = output_format_var.get()
    if fmt == "jpg": safe_run(convert_to_image_jpg)
    elif fmt == "png": safe_run(convert_to_image_png)
    elif fmt == "dxf": safe_run(convert_to_dxf)
    elif engine == "Internal":
        if fmt == "txt": safe_run(extract_text_internal)
        elif fmt == "xlsx": safe_run(convert_to_excel_internal)
        elif fmt == "csv": safe_run(convert_to_csv_internal)
    elif engine == "Tesseract": safe_run(extract_tesseract_task)
    elif engine == "Gemini":
        if not api_key_var.get().strip(): return messagebox.showerror("エラー", "Gemini APIキーを入力してください。")
        safe_run(extract_gemini_task)

class CropSelector:
    def __init__(self, master, pdf_path):
        self.top = tk.Toplevel(master); self.top.title("抽出範囲の選択 (複数選択可)"); self.top.configure(bg=BG_COLOR); self.top.transient(master); self.top.grab_set()
        self.pdf_path = pdf_path; self.zoom = 1.0
        btn_frame = ttk.Frame(self.top, padding=10); btn_frame.pack(fill=tk.X)
        ttk.Button(btn_frame, text="クリア", command=self.clear_rects, style="Warning.TButton").pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="設定して閉じる", command=self.save_and_close, style="Primary.TButton").pack(side=tk.RIGHT, padx=5)
        zoom_frame = ttk.Frame(btn_frame); zoom_frame.pack(side=tk.RIGHT, padx=20)
        ttk.Button(zoom_frame, text="拡大 (+)", command=self.zoom_in, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="縮小 (-)", command=self.zoom_out, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Button(zoom_frame, text="フィット", command=self.zoom_fit, width=8).pack(side=tk.LEFT, padx=2)
        ttk.Label(btn_frame, text="【使い方】ドラッグで抽出範囲を囲みます。", foreground=PRIMARY, font=("Segoe UI", 10, "bold")).pack(side=tk.LEFT, padx=10)

        canvas_frame = ttk.Frame(self.top); canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))
        self.vbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL); self.vbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.hbar = ttk.Scrollbar(canvas_frame, orient=tk.HORIZONTAL); self.hbar.pack(side=tk.BOTTOM, fill=tk.X)
        self.canvas = tk.Canvas(canvas_frame, cursor="cross", bg="white", xscrollcommand=self.hbar.set, yscrollcommand=self.vbar.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.vbar.config(command=self.canvas.yview); self.hbar.config(command=self.canvas.xview)

        self.start_x, self.start_y, self.current_rect, self.rectangles = None, None, None, []
        self.canvas.bind("<ButtonPress-1>", self.on_press); self.canvas.bind("<B1-Motion>", self.on_drag); self.canvas.bind("<ButtonRelease-1>", self.on_release)
        if sys.platform.startswith("win"): self.canvas.bind("<MouseWheel>", self.on_mousewheel)

        try:
            self.doc = fitz.open(pdf_path); self.page = self.doc[0]; self.zoom_fit()
        except Exception as e: self.top.destroy(); raise Exception(f"プレビュー生成失敗: {e}")

        self.top.update_idletasks()
        w, h = int(master.winfo_screenwidth() * 0.8), int(master.winfo_screenheight() * 0.8)
        self.top.geometry(f"{w}x{h}+{(master.winfo_screenwidth()//2)-(w//2)}+{(master.winfo_screenheight()//2)-(h//2)}")

    def draw_image(self):
        mat = fitz.Matrix(self.zoom, self.zoom); pix = self.page.get_pixmap(matrix=mat)
        self.tk_image = ImageTk.PhotoImage(Image.frombytes("RGB", [pix.width, pix.height], pix.samples))
        self.canvas.delete("all"); self.canvas.create_image(0, 0, anchor=tk.NW, image=self.tk_image); self.canvas.config(scrollregion=(0, 0, pix.width, pix.height))
        self.img_w, self.img_h = pix.width, pix.height
        for r in self.rectangles: r['id'] = self.canvas.create_rectangle(r['rx1']*self.img_w, r['ry1']*self.img_h, r['rx2']*self.img_w, r['ry2']*self.img_h, outline="red", width=2)

    def zoom_in(self): self.zoom = min(5.0, self.zoom * 1.2); self.draw_image()
    def zoom_out(self): self.zoom = max(0.2, self.zoom / 1.2); self.draw_image()
    def zoom_fit(self): self.zoom = min(2.0, (self.top.winfo_screenheight() * 0.7) / self.page.rect.height); self.draw_image()
    def on_mousewheel(self, event): self.canvas.yview_scroll(int(-1*(event.delta/120)), "units")
    def on_press(self, event):
        self.start_x, self.start_y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        self.current_rect = self.canvas.create_rectangle(self.start_x, self.start_y, self.start_x, self.start_y, outline="red", width=2, dash=(4, 4))
    def on_drag(self, event): self.canvas.coords(self.current_rect, self.start_x, self.start_y, self.canvas.canvasx(event.x), self.canvas.canvasy(event.y))
    def on_release(self, event):
        end_x, end_y = self.canvas.canvasx(event.x), self.canvas.canvasy(event.y)
        if abs(end_x - self.start_x) > 10 and abs(end_y - self.start_y) > 10:
            self.canvas.itemconfig(self.current_rect, dash=())
            self.rectangles.append({'id': self.current_rect, 'rx1': min(self.start_x, end_x)/self.img_w, 'ry1': min(self.start_y, end_y)/self.img_h, 'rx2': max(self.start_x, end_x)/self.img_w, 'ry2': max(self.start_y, end_y)/self.img_h})
        else: self.canvas.delete(self.current_rect)
    def clear_rects(self):
        for r in self.rectangles: self.canvas.delete(r['id'])
        self.rectangles.clear()
    def save_and_close(self):
        global selected_crop_regions
        selected_crop_regions = [(r['rx1'], r['ry1'], r['rx2'], r['ry2']) for r in self.rectangles]
        btn_select_crop.config(text=f"抽出範囲を選択 (設定済: {len(selected_crop_regions)}か所)" if selected_crop_regions else "抽出範囲を選択")
        self.doc.close(); self.top.destroy()

def open_crop_selector():
    files = selected_files if current_mode == "file" else ([os.path.join(selected_folder, f) for f in os.listdir(selected_folder) if f.lower().endswith((".pdf", ".xlsx", ".csv", ".txt"))] if selected_folder else [])
    pdf_files = [f for f in files if f.lower().endswith('.pdf')]
    if not pdf_files: return messagebox.showinfo("情報", "PDFファイルが選択されていません。")
    try: CropSelector(root, pdf_files[0])
    except Exception as e: messagebox.showerror("エラー", str(e))

def reset_crop_regions():
    global selected_crop_regions; selected_crop_regions = []
    btn_select_crop.config(text="抽出範囲を選択")

def select_files():
    global selected_files, selected_folder, current_mode
    files = filedialog.askopenfilenames(filetypes=[("すべての対応ファイル", "*.pdf;*.xlsx;*.csv;*.txt"), ("PDF", "*.pdf")])
    if files: selected_files, selected_folder, current_mode = list(files), "", "file"; update_ui()

def select_folder():
    global selected_folder, selected_files, current_mode
    folder = filedialog.askdirectory(title="フォルダを選択")
    if folder: selected_folder, selected_files, current_mode = folder, [], "folder"; update_ui()

def select_save_dir():
    global preset_save_dir
    folder = filedialog.askdirectory()
    if folder: preset_save_dir = folder; save_label.config(text=preset_save_dir); save_option.set(2)

def on_save_mode_change():
    global preset_save_dir; preset_save_dir = ""
    save_label.config(text="同じフォルダ" if save_option.get() == 1 else "未選択")

format_radiobuttons = {}
def toggle_extraction_settings(*args):
    is_active = current_mode is not None
    for fmt, rb in format_radiobuttons.items(): rb.configure(state=tk.NORMAL if is_active else tk.DISABLED)
    state_gemini = tk.NORMAL if (is_active and engine_var.get() == "Gemini") else tk.DISABLED
    api_key_entry.configure(state=state_gemini); btn_api_test.configure(state=state_gemini)
    state_crop = tk.NORMAL if is_active else tk.DISABLED
    for child in crop_frame.winfo_children():
        if isinstance(child, ttk.Button) or isinstance(child, ttk.Label): child.configure(state=state_crop)

def update_ui():
    path_label.config(text="\n".join(selected_files) if current_mode == "file" else (f"フォルダ: {selected_folder}" if selected_folder else "未選択"))
    is_active = current_mode is not None
    state_val = tk.NORMAL if is_active else tk.DISABLED
    btn_split.config(state=state_val); btn_rotate.config(state=state_val); btn_extract.config(state=state_val); btn_aggregate.config(state=state_val)
    btn_merge.config(state=tk.NORMAL if current_mode=="folder" else tk.DISABLED)
    if not is_active: reset_crop_regions()
    toggle_extraction_settings()

def show_text_window(title, content):
    win = tk.Toplevel(root); win.title(title); win.geometry("620x550"); win.configure(bg=BG_COLOR)
    x = root.winfo_x() + (WINDOW_WIDTH // 2) - 310; y = root.winfo_y() + (WINDOW_HEIGHT // 2) - 275; win.geometry(f"+{x}+{y}")
    text_area = st.ScrolledText(win, wrap=tk.WORD, font=("Meiryo UI", 10), bg=CARD_BG, fg=TEXT_COLOR, relief=tk.FLAT, padx=15, pady=15)
    text_area.pack(expand=True, fill=tk.BOTH, padx=10, pady=10); text_area.insert(tk.END, content); text_area.configure(state=tk.DISABLED)

# ==============================
# UI画面の構築
# ==============================
root = tk.Tk(); root.title(f"{APP_TITLE} {VERSION}"); root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}+0+0"); root.configure(bg=BG_COLOR)
style = ttk.Style(); style.theme_use("clam") if "clam" in style.theme_names() else None
style.configure(".", background=BG_COLOR, font=("Segoe UI", 10))
style.configure("Card.TFrame", background=CARD_BG)
style.configure("Card.TLabelframe", background=CARD_BG, borderwidth=1, bordercolor=BORDER_COLOR)
style.configure("Card.TLabelframe.Label", background=CARD_BG, foreground=PRIMARY, font=("Segoe UI", 11, "bold"))
style.configure("TButton", padding=6, font=("Segoe UI", 10), background="#E9ECEF", foreground=TEXT_COLOR, borderwidth=1)
style.map("TButton", background=[("active", "#DEE2E6")])
style.configure("Primary.TButton", background=PRIMARY, foreground="white", borderwidth=0)
style.map("Primary.TButton", background=[("active", PRIMARY_HOVER)])
style.configure("Warning.TButton", background=COLOR_WARNING, foreground="black", borderwidth=0)
style.map("Warning.TButton", background=[("active", COLOR_WARNING_HOVER)])
style.configure("Danger.TButton", background=COLOR_DANGER, foreground="white", borderwidth=0)
style.map("Danger.TButton", background=[("active", COLOR_DANGER_HOVER)])
style.configure("Purple.TButton", background=COLOR_PURPLE, foreground="white", borderwidth=0)
style.map("Purple.TButton", background=[("active", COLOR_PURPLE_HOVER)])
style.configure("TRadiobutton", background=CARD_BG, font=("Segoe UI", 10), foreground=TEXT_COLOR)

menubar = Menu(root); help_menu = Menu(menubar, tearoff=0)
help_menu.add_command(label="AI抽出の使い方", command=lambda: show_text_window("AI抽出の使い方", AI_HELP_TEXT.strip()))
menubar.add_cascade(label="ヘルプ", menu=help_menu); root.config(menu=menubar)

rotate_option, save_option = tk.IntVar(value=270), tk.IntVar(value=1)
engine_var, output_format_var, api_key_var = tk.StringVar(value="Internal"), tk.StringVar(value="xlsx"), tk.StringVar(value=get_api_key() or "")
engine_var.trace("w", toggle_extraction_settings)

main_container = ttk.Frame(root, padding=15); main_container.pack(fill=tk.BOTH, expand=True)
title_frame = ttk.Frame(main_container); title_frame.pack(fill=tk.X, pady=(0, 10))
ttk.Label(title_frame, text=APP_TITLE, font=("Segoe UI", 20, "bold"), foreground=PRIMARY).pack(side=tk.LEFT)
ttk.Label(title_frame, text=f" {VERSION}", font=("Segoe UI", 12), foreground=MUTED_TEXT).pack(side=tk.LEFT, pady=(8, 0))

file_card = ttk.Frame(main_container, style="Card.TFrame", padding=10); file_card.pack(fill=tk.X, pady=5)
btn_frame = ttk.Frame(file_card, style="Card.TFrame"); btn_frame.pack()
ttk.Button(btn_frame, text="📄 ファイルを選択", command=select_files, width=22, style="Primary.TButton").grid(row=0, column=0, padx=8)
ttk.Button(btn_frame, text="📁 フォルダを選択", command=select_folder, width=22, style="Primary.TButton").grid(row=0, column=1, padx=8)
path_label = ttk.Label(file_card, text="未選択", background=CARD_BG, foreground=TEXT_COLOR, wraplength=580, justify="center"); path_label.pack(pady=(10, 0))

settings_grid = ttk.Frame(main_container); settings_grid.pack(fill=tk.X, pady=5); settings_grid.columnconfigure(0, weight=1); settings_grid.columnconfigure(1, weight=1)
save_frame = ttk.LabelFrame(settings_grid, text=" 保存先設定 ", style="Card.TLabelframe", padding=8); save_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
ttk.Radiobutton(save_frame, text="元のファイルと同じフォルダ", variable=save_option, value=1, command=on_save_mode_change).pack(anchor="w", pady=2)
ttk.Radiobutton(save_frame, text="任意のフォルダを指定", variable=save_option, value=2, command=on_save_mode_change).pack(anchor="w", pady=2)
ttk.Button(save_frame, text="📂 フォルダ参照", command=select_save_dir).pack(pady=(8, 2))
save_label = ttk.Label(save_frame, text="同じフォルダ", background=CARD_BG, foreground=MUTED_TEXT, font=("Segoe UI", 9)); save_label.pack()

rotate_frame = ttk.LabelFrame(settings_grid, text=" 回転設定 ", style="Card.TLabelframe", padding=8); rotate_frame.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
for t, v in [("左（270°）", 270), ("上下（180°）", 180), ("右（90°）", 90)]: ttk.Radiobutton(rotate_frame, text=t, variable=rotate_option, value=v).pack(anchor="w", pady=4)

extract_frame = ttk.LabelFrame(main_container, text=" ⚙️ データ抽出・変換設定 ", style="Card.TLabelframe", padding=8); extract_frame.pack(fill=tk.X, pady=10)
engine_frame = ttk.Frame(extract_frame, style="Card.TFrame"); engine_frame.pack(fill=tk.X, pady=(0, 5))
ttk.Label(engine_frame, text="① エンジン:", width=12, background=CARD_BG, font=("Segoe UI", 10, "bold"), foreground=TEXT_COLOR).pack(side=tk.LEFT)
engine_inner = ttk.Frame(engine_frame, style="Card.TFrame"); engine_inner.pack(anchor="w", fill=tk.X)
for text, val in [("Python標準ライブラリ (高速・オフライン)", "Internal"), ("Tesseract (ローカルOCR)", "Tesseract"), ("Gemini API (超高精度AI)", "Gemini")]:
    ttk.Radiobutton(engine_inner, text=text, variable=engine_var, value=val).pack(anchor="w", pady=2)

format_frame = ttk.Frame(extract_frame, style="Card.TFrame"); format_frame.pack(fill=tk.X, pady=5)
ttk.Label(format_frame, text="② 出力形式:", width=12, background=CARD_BG, font=("Segoe UI", 10, "bold"), foreground=TEXT_COLOR).pack(side=tk.LEFT)
formats = [("Excel (.xlsx)", "xlsx"), ("CSV (.csv)", "csv"), ("Text (.txt)", "txt"), ("JPEG (.jpg)", "jpg"), ("PNG (.png)", "png"), ("DXF (.dxf)", "dxf")]
format_inner1 = ttk.Frame(format_frame, style="Card.TFrame"); format_inner1.pack(anchor="w", fill=tk.X)
for text, val in formats[:3]:
    rb = ttk.Radiobutton(format_inner1, text=text, variable=output_format_var, value=val); rb.pack(side=tk.LEFT, padx=(0, 15)); format_radiobuttons[val] = rb
format_inner2 = ttk.Frame(format_frame, style="Card.TFrame"); format_inner2.pack(anchor="w", fill=tk.X, pady=(5, 0))
for text, val in formats[3:]:
    rb = ttk.Radiobutton(format_inner2, text=text, variable=output_format_var, value=val); rb.pack(side=tk.LEFT, padx=(0, 15)); format_radiobuttons[val] = rb

ttk.Separator(extract_frame, orient="horizontal").pack(fill=tk.X, pady=8)

api_key_frame = ttk.Frame(extract_frame, style="Card.TFrame"); api_key_frame.pack(fill=tk.X, pady=5)
ttk.Label(api_key_frame, text="[AI用] APIキー:", width=14, background=CARD_BG, font=("Segoe UI", 9, "bold"), foreground=TEXT_COLOR).pack(side=tk.LEFT)
api_key_entry = ttk.Entry(api_key_frame, textvariable=api_key_var, width=45, show="*"); api_key_entry.pack(side=tk.LEFT, padx=(0, 8))
btn_api_test = ttk.Button(api_key_frame, text="テスト", command=test_api_key_ui, width=6); btn_api_test.pack(side=tk.LEFT)

crop_frame = ttk.Frame(extract_frame, style="Card.TFrame"); crop_frame.pack(fill=tk.X, pady=(2, 0))
ttk.Label(crop_frame, text="抽出範囲:", width=14, background=CARD_BG, font=("Segoe UI", 9, "bold"), foreground=TEXT_COLOR).pack(side=tk.LEFT)
btn_select_crop = ttk.Button(crop_frame, text="抽出範囲を選択", command=open_crop_selector); btn_select_crop.pack(side=tk.LEFT)
btn_reset_crop = ttk.Button(crop_frame, text="全体に戻す", command=reset_crop_regions, style="Warning.TButton"); btn_reset_crop.pack(side=tk.LEFT, padx=(5, 0))

action_container = ttk.Frame(main_container); action_container.pack(fill=tk.BOTH, expand=True, pady=5)
action_container.columnconfigure(0, weight=1); action_container.columnconfigure(1, weight=1)

pdf_action_frame = ttk.LabelFrame(action_container, text=" ✂️ PDF編集 ", style="Card.TLabelframe", padding=10); pdf_action_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
btn_merge = ttk.Button(pdf_action_frame, text="結合 (フォルダ)", command=lambda: safe_run(merge_pdfs), style="Primary.TButton"); btn_merge.pack(fill=tk.X, pady=4)
btn_split = ttk.Button(pdf_action_frame, text="分割", command=lambda: safe_run(split_pdfs), style="Primary.TButton"); btn_split.pack(fill=tk.X, pady=4)
btn_rotate = ttk.Button(pdf_action_frame, text="回転", command=lambda: safe_run(rotate_pdfs), style="Primary.TButton"); btn_rotate.pack(fill=tk.X, pady=4)

data_action_frame = ttk.LabelFrame(action_container, text=" 📊 データ操作 ", style="Card.TLabelframe", padding=10); data_action_frame.grid(row=0, column=1, sticky="nsew", padx=(8, 0))
btn_extract = ttk.Button(data_action_frame, text="🚀 選択した抽出・変換を実行", command=run_selected_extraction, style="Danger.TButton"); btn_extract.pack(fill=tk.X, pady=(4, 10), ipady=6) 
btn_aggregate = ttk.Button(data_action_frame, text="🧩 データ集約", command=lambda: safe_run(aggregate_only_task), style="Purple.TButton"); btn_aggregate.pack(fill=tk.X, pady=4)

update_ui()
root.mainloop()