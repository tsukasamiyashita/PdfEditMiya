# -*- coding: utf-8 -*-
import os
import csv
from PyPDF2 import PdfReader, PdfWriter
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Border, Side
import fitz  # PyMuPDF
import ezdxf
import numpy as np
import cv2
from PIL import Image

from utils import auto_adjust_excel_column_width, sanitize_excel_text

# ==============================
# コア処理関数群 (PDF編集・標準抽出)
# ==============================
def merge_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    writer = PdfWriter()
    ui.update_overall(0, len(files), f"全体の進捗 ( 0 / {len(files)} ファイル )")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ファイルを結合中... ( {j} / {len(reader.pages)} ページ )")
            writer.add_page(p)
    ui.set_determinate(1, 1, "PDFを保存中...")
    if save_dir:
        with open(os.path.join(save_dir, f"{options.get('folder_name', 'Merged')}_Merge.pdf"), "wb") as out: writer.write(out)

def split_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(reader.pages))))
        for n, p in enumerate(reader.pages, 1):
            ui.set_determinate(n, len(reader.pages), f"ファイルを分割中... ( {n} / {len(reader.pages)} ページ )")
            writer = PdfWriter(); writer.add_page(p)
            with open(os.path.join(save_dir, f"{base}_Split_{str(n).zfill(digits)}.pdf"), "wb") as out: writer.write(out)

def rotate_pdfs(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        reader = PdfReader(f); writer = PdfWriter()
        for j, p in enumerate(reader.pages, 1):
            ui.set_determinate(j, len(reader.pages), f"ページを回転中... ( {j} / {len(reader.pages)} ページ )")
            p.rotate(options.get("rotate_deg", 270)); writer.add_page(p)
        base = os.path.splitext(os.path.basename(f))[0]
        with open(os.path.join(save_dir, f"{base}_Rotate.pdf"), "wb") as out: writer.write(out)

def extract_text_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(f))[0]; text_list = []
        with pdfplumber.open(f) as pdf:
            for j, p in enumerate(pdf.pages, 1):
                ui.set_determinate(j, len(pdf.pages), f"テキストを抽出中... ( {j} / {len(pdf.pages)} ページ )")
                if crop_regions:
                    page_texts = []
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * p.width, min(ry1, ry2) * p.height, max(rx1, rx2) * p.width, max(ry1, ry2) * p.height
                        txt = p.crop((x0, top, x1, bottom)).extract_text()
                        if txt: page_texts.append(txt)
                    if page_texts: text_list.append("\n".join(page_texts))
                else:
                    txt = p.extract_text()
                    if txt: text_list.append(txt)
        with open(os.path.join(save_dir, f"{base}_Text.txt"), "w", encoding="utf-8") as out: out.write("\n".join(text_list))

def convert_to_excel_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    border_style = Side(border_style="thin", color="000000")
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        wb = Workbook(); wb.remove(wb.active)
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"表データをExcelへ変換中... ( {page_idx} / {len(pdf.pages)} ページ )")
                tables = []
                if crop_regions:
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * page.width, min(ry1, ry2) * page.height, max(rx1, rx2) * page.width, max(ry1, ry2) * page.height
                        tbls = page.crop((x0, top, x1, bottom)).extract_tables()
                        if tbls: tables.extend(tbls)
                else: tables = page.extract_tables()
                if not tables: continue
                ws = wb.create_sheet(f"Page_{str(page_idx).zfill(digits)}"); current_row = 1
                for table in tables:
                    for row_data in table:
                        for col_idx, cell_value in enumerate(row_data, 1):
                            clean_value = sanitize_excel_text(cell_value)
                            cell = ws.cell(row=current_row, column=col_idx, value=clean_value if clean_value else "")
                            cell.border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                        current_row += 1
                    current_row += 2
                auto_adjust_excel_column_width(ws)
        if len(wb.sheetnames) > 0: wb.save(os.path.join(save_dir, f"{os.path.splitext(os.path.basename(pdf_path))[0]}_Excel.xlsx"))

def convert_to_csv_internal(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, pdf_path in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        with pdfplumber.open(pdf_path) as pdf:
            digits = max(2, len(str(len(pdf.pages))))
            for page_idx, page in enumerate(pdf.pages, 1):
                ui.set_determinate(page_idx, len(pdf.pages), f"表データをCSVへ変換中... ( {page_idx} / {len(pdf.pages)} ページ )")
                tables = []
                if crop_regions:
                    for (rx1, ry1, rx2, ry2) in crop_regions:
                        x0, top, x1, bottom = min(rx1, rx2) * page.width, min(ry1, ry2) * page.height, max(rx1, rx2) * page.width, max(ry1, ry2) * page.height
                        tbls = page.crop((x0, top, x1, bottom)).extract_tables()
                        if tbls: tables.extend(tbls)
                else: tables = page.extract_tables()
                if not tables: continue
                with open(os.path.join(save_dir, f"{base}_Page_{str(page_idx).zfill(digits)}_CSV.csv"), "w", encoding="utf-8-sig", newline="") as f_out:
                    writer = csv.writer(f_out)
                    for table in tables:
                        for row_data in table: writer.writerow([str(cell).strip() if cell else "" for cell in row_data])
                        writer.writerow([]) 

def convert_to_image_jpg(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "jpg")
def convert_to_image_png(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "png")
def convert_to_image_tiff(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "tiff")
def convert_to_image_bmp(files, save_dir, options, ui): _convert_image(files, save_dir, options, ui, "bmp")

def _convert_image(files, save_dir, options, ui, ext):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        doc = fitz.open(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(doc))))
        for n, page in enumerate(doc, 1):
            ui.set_determinate(n, len(doc), f"画像へ変換中... ( {n} / {len(doc)} ページ )")
            n_str = str(n).zfill(digits)
            pix = page.get_pixmap(dpi=200)
            if crop_regions:
                img_array = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                if pix.n == 4: img_array = cv2.cvtColor(img_array, cv2.COLOR_RGBA2RGB)
                elif pix.n == 1: img_array = cv2.cvtColor(img_array, cv2.COLOR_GRAY2RGB)
                h, w = img_array.shape[:2]
                for idx, (rx1, ry1, rx2, ry2) in enumerate(crop_regions):
                    x1, y1 = int(min(rx1, rx2) * w), int(min(ry1, ry2) * h)
                    x2, y2 = int(max(rx1, rx2) * w), int(max(ry1, ry2) * h)
                    Image.fromarray(img_array[y1:y2, x1:x2]).save(os.path.join(save_dir, f"{base}_{n_str}_crop{idx+1}.{ext}"))
            else: 
                # PNG, JPG 以外は PIL を経由して保存
                if ext in ["tiff", "bmp"]:
                    Image.frombytes("RGB" if pix.n >= 3 else "L", [pix.width, pix.height], pix.samples).save(os.path.join(save_dir, f"{base}_{n_str}.{ext}"))
                else:
                    pix.save(os.path.join(save_dir, f"{base}_{n_str}.{ext}"))

def convert_to_svg(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        doc = fitz.open(f)
        base = os.path.splitext(os.path.basename(f))[0]
        digits = max(2, len(str(len(doc))))
        for n, page in enumerate(doc, 1):
            ui.set_determinate(n, len(doc), f"SVGへ変換中... ( {n} / {len(doc)} ページ )")
            n_str = str(n).zfill(digits)
            if crop_regions:
                h, w = page.rect.height, page.rect.width
                for idx, (rx1, ry1, rx2, ry2) in enumerate(crop_regions):
                    rect = fitz.Rect(min(rx1, rx2)*w, min(ry1, ry2)*h, max(rx1, rx2)*w, max(ry1, ry2)*h)
                    page.set_cropbox(rect)
                    svg_xml = page.get_svg_image()
                    with open(os.path.join(save_dir, f"{base}_{n_str}_crop{idx+1}.svg"), "w", encoding="utf-8") as svg_file:
                        svg_file.write(svg_xml)
                    page.set_cropbox(page.mediabox) # 元に戻す
            else:
                svg_xml = page.get_svg_image()
                with open(os.path.join(save_dir, f"{base}_{n_str}.svg"), "w", encoding="utf-8") as svg_file:
                    svg_file.write(svg_xml)

def convert_to_dxf(files, save_dir, options, ui):
    files = [f for f in files if f.lower().endswith(".pdf")]
    if not files: raise Exception("PDFファイルが含まれていません。")
    crop_regions = options.get("crop_regions", [])
    def in_crop(x, y, w, h):
        if not crop_regions: return True
        for (rx1, ry1, rx2, ry2) in crop_regions:
            if min(rx1, rx2)*w <= x <= max(rx1, rx2)*w and min(ry1, ry2)*h <= y <= max(ry1, ry2)*h: return True
        return False

    for i, f in enumerate(files, 1):
        ui.update_overall(i, len(files), f"全体の進捗 ( {i} / {len(files)} ファイル )")
        try:
            doc = fitz.open(f); dwg = ezdxf.new('R2010'); msp = dwg.modelspace()
            for page_num, page in enumerate(doc, 1):
                ui.set_determinate(page_num, len(doc), f"DXFへ変換中... ( {page_num} / {len(doc)} ページ )")
                h, w = page.rect.height, page.rect.width
                paths = page.get_drawings()
                is_vector_rich = len(paths) > 0
                if is_vector_rich:
                    for path in paths:
                        for item in path["items"]:
                            if item[0] == "l":
                                p1, p2 = item[1], item[2]
                                if in_crop(p1.x, p1.y, w, h) or in_crop(p2.x, p2.y, w, h): msp.add_line((p1.x, h - p1.y), (p2.x, h - p2.y))
                            elif item[0] == "re":
                                rect = item[1]
                                if in_crop(rect.x0, rect.y0, w, h) or in_crop(rect.x1, rect.y1, w, h):
                                    msp.add_lwpolyline([(rect.x0, h - rect.y0), (rect.x1, h - rect.y0), (rect.x1, h - rect.y1), (rect.x0, h - rect.y1)], close=True)
                            elif item[0] == "c":
                                p1, p2, p3, p4 = item[1], item[2], item[3], item[4]
                                if any(in_crop(pt.x, pt.y, w, h) for pt in [p1, p2, p3, p4]):
                                    msp.add_spline([(p1.x, h - p1.y), (p2.x, h - p2.y), (p3.x, h - p3.y), (p4.x, h - p4.y)])
                if not is_vector_rich or len(paths) < 5:
                    pix = page.get_pixmap(dpi=300)
                    img = np.frombuffer(pix.samples, dtype=np.uint8).reshape(pix.h, pix.w, pix.n)
                    gray = cv2.cvtColor(img, cv2.COLOR_RGB2GRAY) if pix.n >= 3 else img
                    if crop_regions:
                        mask = np.zeros_like(gray)
                        for (rx1, ry1, rx2, ry2) in crop_regions:
                            mask[int(min(ry1, ry2)*gray.shape[0]):int(max(ry1, ry2)*gray.shape[0]), int(min(rx1, rx2)*gray.shape[1]):int(max(rx1, rx2)*gray.shape[1])] = 255
                        gray = cv2.bitwise_and(gray, mask)
                    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
                    binary = cv2.morphologyEx(cv2.adaptiveThreshold(blurred, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY_INV, 11, 2), cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
                    contours, _ = cv2.findContours(binary, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
                    scale_x, scale_y = w / pix.w, h / pix.h
                    for cnt in contours:
                        if cv2.contourArea(cnt) < 15: continue 
                        pts = [(p[0][0] * scale_x, h - p[0][1] * scale_y) for p in cv2.approxPolyDP(cnt, 0.003 * cv2.arcLength(cnt, True), True)]
                        if len(pts) > 1: msp.add_lwpolyline(pts, close=True)
            ui.set_determinate(len(doc), len(doc), "DXFファイルを保存中...")
            dwg.saveas(os.path.join(save_dir, f"{os.path.splitext(os.path.basename(f))[0]}_CAD.dxf"))
        except Exception as e: print(f"DXF Conversion Error: {e}")